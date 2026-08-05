"""
recon.py — 패시브 정보수집(OSINT) 정찰 모듈
==============================================================================
탐지 모듈(scan() 인터페이스)·SQLi 추출·엑셀 취합과 완전히 분리된 별도 모드.

**하드 룰: 대상 도메인·서브도메인·서버로는 어떤 요청도 직접 보내지 않는다
(순수 패시브).** 이 모듈이 직접 접속하는 호스트는 아래 6개뿐이다.

  - crt.sh                 : CT(Certificate Transparency) 로그 조회
  - web.archive.org        : Wayback Machine CDX 인덱스 조회 + robots.txt/sitemap.xml
                             아카이브 스냅샷 본문 조회 (대상이 아닌 아카이브에서 읽음 —
                             대상 서버로는 요청이 가지 않음)
  - index.commoncrawl.org  : Common Crawl 인덱스(CDX) 조회 (무키)
  - urlscan.io             : 기존 공개 스캔 결과 검색 (search API, 무키·읽기 전용 —
                             스캔 제출 엔드포인트(/scan)는 호출하지 않음)
  - 8.8.8.8 / 1.1.1.1      : 공용 DNS 리졸버 (대상 네임서버로 직접 질의하지 않음 —
                             캐시 미스 시 재귀 질의는 리졸버가 대신 수행)
  - internetdb.shodan.io   : Shodan이 사전 수집한 IP별 포트 정보 조회
                             (무키, 읽기 전용 — 온디맨드 스캔 API 아님)

이 6개 외의 호스트로 requests가 나가는 코드 경로는 존재하지 않는다.
"""
import html as html_module
import ipaddress
import json
import os
import re
import xml.etree.ElementTree as ET
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Set, Tuple
from urllib.parse import urlparse, urljoin

import requests

from modules._cancel import wait_or_cancel

# ── 상수 ────────────────────────────────────────────────────────────────────

CRTSH_URL = "https://crt.sh/"
WAYBACK_CDX_URL = "http://web.archive.org/cdx/search/cdx"
WAYBACK_AVAILABLE_URL = "http://archive.org/wayback/available"
COMMONCRAWL_COLLINFO_URL = "https://index.commoncrawl.org/collinfo.json"
# collinfo.json 조회 실패 시에만 쓰는 최후 fallback 인덱스 id (주기적 갱신 필요)
COMMONCRAWL_FALLBACK_INDEXES = ("CC-MAIN-2026-25", "CC-MAIN-2026-18", "CC-MAIN-2026-10")
URLSCAN_SEARCH_URL = "https://urlscan.io/api/v1/search/"
INTERNETDB_URL = "https://internetdb.shodan.io/{ip}"

# 지원 소스 키 — 대시보드 토글과 1:1 매핑
SOURCE_KEYS = ("crtsh", "wayback", "dns", "commoncrawl", "urlscan", "archivepaths", "internetdb")

# 공용 DNS 리졸버 — 대상 네임서버 직접 질의 금지, 이 리졸버로만 조회
PUBLIC_RESOLVERS = ["8.8.8.8", "1.1.1.1"]

# apex(입력) 도메인 전용 — 전체 레코드 타입 조회
APEX_RECORD_TYPES = ("A", "AAAA", "MX", "NS", "TXT", "SOA", "CAA")
# 서브도메인 전용 — 경량 레코드 타입만 조회 (질의량 절감, 서브도메인은
# 보통 자체 SOA/NS/MX를 갖지 않음)
SUBDOMAIN_RECORD_TYPES = ("A", "AAAA", "CNAME")

# 서브도메인 DNS 확인 상한 기본값/범위
DEFAULT_MAX_SUBDOMAINS = 200
MIN_MAX_SUBDOMAINS = 10
MAX_MAX_SUBDOMAINS = 1000

# Common Crawl 최근 인덱스 조회 개수
CC_INDEX_COUNT = 3
# 아카이브 robots.txt/sitemap.xml 파싱 대상 서브도메인 상한 (apex 우선)
MAX_ROBOTS_HOSTS = 25
# sitemap.xml/사이트맵 인덱스 하위 fetch 총 상한 (1단계 재귀)
MAX_SITEMAP_FETCH = 10
# 소스별 원시 URL 수집 상한 (그룹핑 전 1차 방어 — 과다 팽창 방지)
MAX_RAW_URLS_PER_SOURCE = 5000
# 서브도메인별 / 전체 URL 최종 노출 상한
MAX_URLS_PER_HOST = 200
MAX_TOTAL_URLS = 3000

# 엑셀 수식 인젝션 방어 접두사
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# 도메인 형식 검증 — 호스트명만 허용 (스킴·경로·공백·단일 라벨 불가)
_DOMAIN_RE = re.compile(
    r"^(?!-)[A-Za-z0-9-]{1,63}(?<!-)(\.(?!-)[A-Za-z0-9-]{1,63}(?<!-))+$"
)


# ── 도메인 검증 ──────────────────────────────────────────────────────────────

def normalize_domain(raw: str) -> str:
    """사용자 입력을 도메인 문자열로 정규화한다.

    "https://example.com/path" 형태로 붙여넣어도 스킴·포트·경로를 제거하고
    호스트명만 추출한다. 형식이 올바르지 않으면 ValueError.
    """
    raw = (raw or "").strip()
    if "://" in raw:
        raw = urlparse(raw).netloc or raw
    raw = raw.split("/")[0].split(":")[0].strip().rstrip(".").lower()
    if not _DOMAIN_RE.match(raw):
        raise ValueError(f"올바르지 않은 도메인 형식: {raw!r}")
    return raw


def _is_in_scope(host: str, domain: str) -> bool:
    """host가 domain 자신이거나 domain의 서브도메인인지 확인 (무관 도메인 오염 방지)."""
    host = host.lower().rstrip(".")
    domain = domain.lower().rstrip(".")
    return host == domain or host.endswith("." + domain)


def _ip_sort_key(ip: str):
    """IPv4/IPv6 혼합 목록을 안정적으로 정렬하기 위한 키."""
    try:
        addr = ipaddress.ip_address(ip)
        return (addr.version, addr.packed)
    except ValueError:
        return (99, ip.encode())


# ── 소스 1: crt.sh (CT 로그) ─────────────────────────────────────────────────

def query_crtsh(domain: str, timeout: int, session: requests.Session) -> Dict[str, Any]:
    """crt.sh CT 로그 조회 — 서브도메인 + 인증서 메타 반환.

    실패해도 예외를 올리지 않고 error 필드가 채워진 빈 결과를 반환한다
    (다른 소스는 계속 진행되어야 하므로).
    """
    subdomains: Set[str] = set()
    certificates: List[Dict[str, Any]] = []
    try:
        resp = session.get(
            CRTSH_URL, params={"q": f"%.{domain}", "output": "json"},
            timeout=timeout,
        )
        resp.raise_for_status()
        entries = resp.json()

        seen_cert_ids: Set[Any] = set()
        for entry in entries:
            name_value = entry.get("name_value", "") or ""
            for name in name_value.split("\n"):
                name = name.strip().lower().lstrip("*.")
                if name and _is_in_scope(name, domain):
                    subdomains.add(name)
            cert_id = entry.get("id")
            if cert_id is not None and cert_id not in seen_cert_ids:
                seen_cert_ids.add(cert_id)
                certificates.append({
                    "id": cert_id,
                    "common_name": entry.get("common_name", ""),
                    "issuer": entry.get("issuer_name", ""),
                    "not_before": entry.get("not_before", ""),
                    "not_after": entry.get("not_after", ""),
                })
    except Exception as e:
        return {"subdomains": subdomains, "certificates": certificates,
                "error": f"crt.sh 조회 실패: {e}"}
    return {"subdomains": subdomains, "certificates": certificates, "error": None}


# ── 소스 2: Wayback Machine CDX 인덱스 ───────────────────────────────────────

def query_wayback(domain: str, timeout: int, session: requests.Session) -> Dict[str, Any]:
    """Wayback Machine CDX 인덱스 조회 — 서브도메인 + 아카이브 URL 반환.

    matchType=domain으로 대상 도메인 + 전체 서브도메인의 아카이브 URL을 한 번에
    조회한다. 스냅샷 본문은 조회하지 않고 CDX 인덱스(URL 목록)만 읽는다.
    """
    subdomains: Set[str] = set()
    urls: List[str] = []
    try:
        resp = session.get(
            WAYBACK_CDX_URL,
            params={
                "url": domain,
                "matchType": "domain",
                "output": "json",
                "fl": "original",
                "collapse": "urlkey",
                "limit": 10000,
            },
            timeout=timeout,
        )
        resp.raise_for_status()
        rows = resp.json()
        # 첫 행은 헤더(["original"]) — 데이터 없으면 rows가 빈 리스트
        for row in (rows[1:] if len(rows) > 1 else []):
            original = row[0] if row else ""
            if not original:
                continue
            host = urlparse(original).netloc.split(":")[0].lower()
            if host and _is_in_scope(host, domain):
                subdomains.add(host)
            if len(urls) < MAX_RAW_URLS_PER_SOURCE:
                urls.append(original)
    except Exception as e:
        return {"subdomains": subdomains, "urls": urls,
                "error": f"Wayback 조회 실패: {e}"}
    return {"subdomains": subdomains, "urls": urls, "error": None}


# ── 소스 3: Common Crawl 인덱스 ──────────────────────────────────────────────

def _commoncrawl_indexes(timeout: int, session: requests.Session) -> List[Tuple[str, str]]:
    """collinfo.json에서 최신 CC_INDEX_COUNT개 (id, cdx-api URL) 목록을 확보한다.

    조회 실패 시 하드코딩된 COMMONCRAWL_FALLBACK_INDEXES로 대체한다(최후 수단).
    """
    try:
        resp = session.get(COMMONCRAWL_COLLINFO_URL, timeout=timeout)
        resp.raise_for_status()
        collinfo = resp.json()
        pairs = [(c["id"], c["cdx-api"]) for c in collinfo if c.get("id") and c.get("cdx-api")]
        if pairs:
            return pairs[:CC_INDEX_COUNT]
    except Exception:
        pass
    return [(cid, f"https://index.commoncrawl.org/{cid}-index")
            for cid in COMMONCRAWL_FALLBACK_INDEXES[:CC_INDEX_COUNT]]


def query_commoncrawl(domain: str, timeout: int, session: requests.Session) -> Dict[str, Any]:
    """Common Crawl 인덱스 조회 — 서브도메인 + 관측 URL 반환.

    최근 CC_INDEX_COUNT개 크롤 인덱스를 matchType=domain으로 각각 조회한다
    (JSONL 응답 — Wayback CDX와 달리 헤더 행이 없다). 개별 인덱스 조회 실패는
    건너뛰고 나머지 인덱스로 계속 진행한다.
    """
    subdomains: Set[str] = set()
    urls: List[str] = []
    seen: Set[str] = set()
    failed: List[str] = []

    for index_id, cdx_api in _commoncrawl_indexes(timeout, session):
        try:
            resp = session.get(
                cdx_api,
                params={"url": domain, "matchType": "domain", "output": "json", "limit": 2000},
                timeout=timeout,
            )
            resp.raise_for_status()
            for line in resp.text.splitlines():
                line = line.strip()
                if not line:
                    continue
                try:
                    row = json.loads(line)
                except ValueError:
                    continue
                original = row.get("url", "")
                if not original or original in seen:
                    continue
                seen.add(original)
                host = urlparse(original).netloc.split(":")[0].lower()
                if host and _is_in_scope(host, domain):
                    subdomains.add(host)
                if len(urls) < MAX_RAW_URLS_PER_SOURCE:
                    urls.append(original)
        except Exception:
            failed.append(index_id)

    error = f"Common Crawl 일부 인덱스 조회 실패: {', '.join(failed)}" if failed else None
    return {"subdomains": subdomains, "urls": urls, "error": error}


# ── 소스 4: urlscan.io (무키 search API) ─────────────────────────────────────

def query_urlscan(domain: str, timeout: int, session: requests.Session) -> Dict[str, Any]:
    """urlscan.io search API(무키) 조회 — 서브도메인 + 관측 URL 반환.

    기존 공개 스캔 결과를 검색(search)만 하는 읽기 전용 조회다 — 스캔 제출
    엔드포인트(/scan)는 호출하지 않는다(대상에 새 요청을 유발하지 않기 위함).
    """
    subdomains: Set[str] = set()
    urls: List[str] = []
    try:
        resp = session.get(
            URLSCAN_SEARCH_URL,
            params={"q": f"domain:{domain}", "size": 100},
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        for entry in data.get("results", []):
            original = (entry.get("page") or {}).get("url", "")
            if not original:
                continue
            host = urlparse(original).netloc.split(":")[0].lower()
            if host and _is_in_scope(host, domain):
                subdomains.add(host)
                if len(urls) < MAX_RAW_URLS_PER_SOURCE:
                    urls.append(original)
    except Exception as e:
        return {"subdomains": subdomains, "urls": urls, "error": f"urlscan.io 조회 실패: {e}"}
    return {"subdomains": subdomains, "urls": urls, "error": None}


# ── 소스 5: 아카이브된 robots.txt / sitemap.xml 파싱 ─────────────────────────

_ROBOTS_PATH_RE = re.compile(r"^(?:disallow|allow)\s*:\s*(\S+)", re.IGNORECASE)
_ROBOTS_SITEMAP_RE = re.compile(r"^sitemap\s*:\s*(\S+)", re.IGNORECASE)


def _wayback_snapshot_body(query_url: str, timeout: int, session: requests.Session) -> Optional[str]:
    """Wayback availability API로 query_url의 최신 스냅샷을 찾아 raw 본문을 반환한다.

    web.archive.org에서만 조회한다 — 대상 서버로는 어떤 요청도 가지 않는다.
    스냅샷이 없거나 조회 실패 시 None(정상 상황 — robots.txt/sitemap.xml이
    아카이브에 없는 호스트가 대부분이므로 예외로 취급하지 않는다).
    """
    try:
        resp = session.get(WAYBACK_AVAILABLE_URL, params={"url": query_url}, timeout=timeout)
        resp.raise_for_status()
        closest = (resp.json().get("archived_snapshots") or {}).get("closest")
        if not closest or not closest.get("available") or not closest.get("url"):
            return None
        # closest["url"] = ".../web/{timestamp}/{archived_original_url}"
        # timestamp 뒤에 "id_"를 삽입하면 Wayback 툴바 없는 raw 본문을 받는다.
        raw_url = re.sub(r"(/web/\d+)/", r"\1id_/", closest["url"], count=1)
        body_resp = session.get(raw_url, timeout=timeout)
        body_resp.raise_for_status()
        return body_resp.text
    except Exception:
        return None


def _parse_robots(text: str, base_url: str) -> Tuple[List[str], List[str]]:
    """robots.txt 본문에서 Disallow/Allow 경로와 Sitemap 지시문을 분리해 추출한다.

    반환: (paths, sitemaps). Disallow/Allow 경로 문자열에 "sitemap"이 우연히
    포함되어도 오분류되지 않도록, 지시문 종류(Sitemap: vs Disallow/Allow:)로만
    구분한다 — 파싱 시점에 이미 분리되므로 호출부에서 문자열 추측이 불필요하다.
    """
    paths: List[str] = []
    sitemaps: List[str] = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        m = _ROBOTS_SITEMAP_RE.match(line)
        if m:
            sitemaps.append(urljoin(base_url, m.group(1).strip()))
            continue
        m = _ROBOTS_PATH_RE.match(line)
        if m:
            path = m.group(1).strip()
            if path and path != "/":
                paths.append(urljoin(base_url, path))
    return paths, sitemaps


def _parse_sitemap(text: str) -> Tuple[List[str], List[str]]:
    """sitemap.xml 본문에서 <loc> URL을 추출한다.

    반환: (urls, sub_sitemaps). 사이트맵 인덱스(<sitemapindex>)면 urls는 비우고
    하위 사이트맵 목록을 sub_sitemaps로 반환한다.
    """
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return [], []
    locs = [el.text.strip() for el in root.iter()
            if el.tag.lower().endswith("loc") and el.text and el.text.strip()]
    if "sitemapindex" in root.tag.lower():
        return [], locs
    return locs, []


def fetch_archive_paths(hosts: List[str], domain: str, timeout: int,
                        session: requests.Session, stop_event=None) -> List[Tuple[str, str]]:
    """대상 서브도메인의 robots.txt/sitemap.xml **아카이브 스냅샷 본문**(web.archive.org)을
    읽어 선언된 엔드포인트 경로를 추출한다. 대상에는 어떤 요청도 보내지 않는다.

    hosts 중 apex(domain) 우선으로 정렬 후 MAX_ROBOTS_HOSTS개까지만 조회한다.
    반환: [(url, "robots"|"sitemap"), ...]
    """
    results: List[Tuple[str, str]] = []
    sitemap_budget = {"n": MAX_SITEMAP_FETCH}
    priority_hosts = sorted(hosts, key=lambda h: (h != domain, h))[:MAX_ROBOTS_HOSTS]

    def _fetch_sitemap(sm_url: str) -> None:
        if sitemap_budget["n"] <= 0 or len(results) >= MAX_RAW_URLS_PER_SOURCE:
            return
        wait_or_cancel(stop_event, 0)
        sitemap_budget["n"] -= 1
        body = _wayback_snapshot_body(sm_url, timeout, session)
        if not body:
            return
        locs, sub_sitemaps = _parse_sitemap(body)
        for loc in locs:
            results.append((loc, "sitemap"))
        for sub in sub_sitemaps:
            if sitemap_budget["n"] <= 0:
                break
            wait_or_cancel(stop_event, 0)
            sitemap_budget["n"] -= 1
            sub_body = _wayback_snapshot_body(sub, timeout, session)
            if sub_body:
                sub_locs, _ = _parse_sitemap(sub_body)
                for loc in sub_locs:
                    results.append((loc, "sitemap"))

    for host in priority_hosts:
        if len(results) >= MAX_RAW_URLS_PER_SOURCE:
            break
        wait_or_cancel(stop_event, 0)
        base_url = f"http://{host}/"
        sitemap_candidates = [base_url + "sitemap.xml"]

        robots_body = _wayback_snapshot_body(f"{host}/robots.txt", timeout, session)
        if robots_body:
            robots_paths, robots_sitemaps = _parse_robots(robots_body, base_url)
            for path in robots_paths:
                results.append((path, "robots"))
            for sm in robots_sitemaps:
                if sm not in sitemap_candidates:
                    sitemap_candidates.append(sm)

        for sm_url in sitemap_candidates:
            _fetch_sitemap(sm_url)

    return results[:MAX_RAW_URLS_PER_SOURCE]


# ── 소스 6: 공용 DNS 리졸버 ───────────────────────────────────────────────────

def _make_resolver(timeout: int):
    """공용 리졸버(PUBLIC_RESOLVERS) 전용 dns.resolver.Resolver를 생성한다.

    OS 기본 리졸버·대상 네임서버 직접 질의를 원천 차단하기 위해
    configure=False로 시스템 설정을 무시하고 nameservers를 직접 지정한다.
    """
    import dns.resolver
    resolver = dns.resolver.Resolver(configure=False)
    resolver.nameservers = list(PUBLIC_RESOLVERS)
    resolver.timeout = min(max(timeout, 1), 3)
    resolver.lifetime = min(max(timeout, 1), 5)
    return resolver


def resolve_dns(host: str, resolver, record_types: Tuple[str, ...]) -> Dict[str, List[str]]:
    """공용 리졸버로 host의 지정 레코드 타입을 조회한다.

    NXDOMAIN/NoAnswer/Timeout 등은 정상적으로 발생할 수 있는 상황이므로
    해당 레코드 타입만 결과에서 생략하고 다음 타입으로 계속 진행한다.
    """
    import dns.exception
    import dns.resolver

    records: Dict[str, List[str]] = {}
    for rtype in record_types:
        try:
            answer = resolver.resolve(host, rtype)
            values = [str(r).rstrip(".") for r in answer]
            if values:
                records[rtype] = values
        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer,
                dns.resolver.NoNameservers, dns.exception.Timeout):
            continue
        except Exception:
            continue
    return records


# ── 소스 7: Shodan InternetDB (무키, 읽기 전용) ──────────────────────────────

def query_internetdb(ip: str, timeout: int, session: requests.Session) -> Optional[Dict[str, Any]]:
    """Shodan InternetDB 무키 조회 — 해당 IP에 대해 Shodan이 사전에 수집해 둔
    데이터를 읽기만 한다(온디맨드 스캔 아님). 데이터 없음(404)이면 None 반환.
    """
    try:
        resp = session.get(INTERNETDB_URL.format(ip=ip), timeout=timeout)
        if resp.status_code == 404:
            return None
        resp.raise_for_status()
        data = resp.json()
        return {
            "ip": data.get("ip", ip),
            "ports": data.get("ports", []),
            "hostnames": data.get("hostnames", []),
            "cpes": data.get("cpes", []),
            "tags": data.get("tags", []),
            "vulns": data.get("vulns", []),
        }
    except Exception:
        return None


# ── 오케스트레이션 ────────────────────────────────────────────────────────────

def run_recon(domain: str, sources: List[str], *,
             timeout: int = 8,
             max_subdomains: int = DEFAULT_MAX_SUBDOMAINS,
             progress_cb: Optional[Callable[[int, int], None]] = None,
             stop_event=None) -> Dict[str, Any]:
    """패시브 정보수집 오케스트레이션.

    sources: SOURCE_KEYS 부분집합. "internetdb"가 있으면 IP 확보를 위해 "dns"를
    방어적으로 자동 포함한다(서버가 이미 보장하지만 직접 호출 대비 재확인).
    progress_cb는 (current, total=100) 형식의 백분율 콜백 — 단계 경계마다 호출된다.
    stop_event가 set되면 각 단계 반복 지점에서 ScanCancelled를 던져 즉시 중단한다.
    """
    sources_set = set(sources) & set(SOURCE_KEYS)
    if "internetdb" in sources_set:
        sources_set.add("dns")

    def _report(pct: int) -> None:
        if progress_cb:
            progress_cb(pct, 100)

    started_at = datetime.now().isoformat()
    errors: List[Dict[str, str]] = []
    session = requests.Session()
    session.headers.update({"User-Agent": "SpaceScan-Recon/1.0 (passive-osint)"})

    all_subdomains: Set[str] = {domain}
    origin: Dict[str, Set[str]] = {domain: set()}
    certificates: List[Dict[str, Any]] = []
    url_sources: Dict[str, Set[str]] = {}

    def _add_origin(hosts: Set[str], source: str) -> None:
        for h in hosts:
            origin.setdefault(h, set()).add(source)

    def _add_urls(urls: List[str], source: str) -> None:
        for u in urls:
            url_sources.setdefault(u, set()).add(source)

    # ── 1단계: 서브도메인 발굴 (crt.sh) ──────────────────────────────────────
    if "crtsh" in sources_set:
        wait_or_cancel(stop_event, 0)
        r = query_crtsh(domain, timeout, session)
        _add_origin(r["subdomains"], "crtsh")
        all_subdomains |= r["subdomains"]
        certificates = r["certificates"]
        if r.get("error"):
            errors.append({"source": "crtsh", "message": r["error"]})
    _report(6)

    # ── 2단계: 서브도메인 발굴 + URL 수집 (Wayback CDX) ─────────────────────
    if "wayback" in sources_set:
        wait_or_cancel(stop_event, 0)
        r = query_wayback(domain, timeout, session)
        _add_origin(r["subdomains"], "wayback")
        all_subdomains |= r["subdomains"]
        _add_urls(r["urls"], "wayback")
        if r.get("error"):
            errors.append({"source": "wayback", "message": r["error"]})
    _report(12)

    # ── 3단계: 서브도메인 발굴 + URL 수집 (Common Crawl) ─────────────────────
    if "commoncrawl" in sources_set:
        wait_or_cancel(stop_event, 0)
        r = query_commoncrawl(domain, timeout, session)
        _add_origin(r["subdomains"], "commoncrawl")
        all_subdomains |= r["subdomains"]
        _add_urls(r["urls"], "commoncrawl")
        if r.get("error"):
            errors.append({"source": "commoncrawl", "message": r["error"]})
    _report(20)

    # ── 4단계: 서브도메인 발굴 + URL 수집 (urlscan.io) ───────────────────────
    if "urlscan" in sources_set:
        wait_or_cancel(stop_event, 0)
        r = query_urlscan(domain, timeout, session)
        _add_origin(r["subdomains"], "urlscan")
        all_subdomains |= r["subdomains"]
        _add_urls(r["urls"], "urlscan")
        if r.get("error"):
            errors.append({"source": "urlscan", "message": r["error"]})
    _report(26)

    # ── 5단계: 아카이브 robots.txt/sitemap.xml 파싱 ──────────────────────────
    if "archivepaths" in sources_set:
        wait_or_cancel(stop_event, 0)
        pairs = fetch_archive_paths(sorted(all_subdomains), domain, timeout, session, stop_event)
        for url, src in pairs:
            host = urlparse(url).netloc.split(":")[0].lower()
            if host and _is_in_scope(host, domain):
                url_sources.setdefault(url, set()).add(src)
    _report(45)

    # ── 6단계: DNS 레코드 조회 (공용 리졸버) ────────────────────────────────
    dns_records: Dict[str, Dict[str, List[str]]] = {}
    resolved_ips: Set[str] = set()
    target_hosts = sorted(all_subdomains)
    truncated = len(target_hosts) > max_subdomains
    if truncated:
        target_hosts = target_hosts[:max_subdomains]

    if "dns" in sources_set and target_hosts:
        resolver = _make_resolver(timeout)
        total = len(target_hosts)
        for i, host in enumerate(target_hosts):
            wait_or_cancel(stop_event, 0)
            rtypes = APEX_RECORD_TYPES if host == domain else SUBDOMAIN_RECORD_TYPES
            recs = resolve_dns(host, resolver, rtypes)
            if recs:
                dns_records[host] = recs
                for rtype in ("A", "AAAA"):
                    resolved_ips.update(recs.get(rtype, []))
            _report(45 + int((i + 1) / total * 40))
    else:
        _report(85)

    # ── 7단계: 열린 포트 조회 (Shodan InternetDB, 무키) ─────────────────────
    ports: Dict[str, Any] = {}
    if "internetdb" in sources_set and resolved_ips:
        ip_list = sorted(resolved_ips, key=_ip_sort_key)
        total = len(ip_list)
        for i, ip in enumerate(ip_list):
            wait_or_cancel(stop_event, 0)
            info = query_internetdb(ip, timeout, session)
            if info:
                ports[ip] = info
            _report(85 + int((i + 1) / total * 15))
    else:
        _report(100)

    finished_at = datetime.now().isoformat()
    subdomains_out = [
        {"host": h, "alive": h in dns_records, "sources": sorted(origin.get(h, set()))}
        for h in sorted(all_subdomains)
    ]

    # ── URL을 서브도메인별로 그룹핑 (스코프 필터 + 상한 적용) ────────────────
    subdomain_urls: Dict[str, List[Dict[str, Any]]] = {}
    url_total = 0
    url_truncated = False
    for url in sorted(url_sources.keys()):
        host = urlparse(url).netloc.split(":")[0].lower()
        if not host or not _is_in_scope(host, domain):
            continue
        if url_total >= MAX_TOTAL_URLS:
            url_truncated = True
            break
        bucket = subdomain_urls.setdefault(host, [])
        if len(bucket) >= MAX_URLS_PER_HOST:
            url_truncated = True
            continue
        bucket.append({"url": url, "sources": sorted(url_sources[url])})
        url_total += 1

    return {
        "domain": domain,
        "subdomains": subdomains_out,
        "dns_records": dns_records,
        "certificates": certificates,
        "subdomain_urls": subdomain_urls,
        "ports": ports,
        "errors": errors,
        "meta": {
            "sources_used": sorted(sources_set),
            "started_at": started_at,
            "finished_at": finished_at,
            "subdomain_total": len(all_subdomains),
            "subdomain_resolved": len(target_hosts) if "dns" in sources_set else 0,
            "subdomain_truncated": truncated,
            "max_subdomains": max_subdomains,
            "resolved_ip_count": len(resolved_ips),
            "url_total": url_total,
            "url_truncated": url_truncated,
        },
    }


# ── 결과 저장: HTML 리포트 ────────────────────────────────────────────────────

_RECON_CSS = """
*{margin:0;padding:0;box-sizing:border-box}
body{background:#070b14;color:#cdd9f0;font-family:'Pretendard',sans-serif;font-size:14px;line-height:1.6}
.header{background:linear-gradient(135deg,#0f1628,#1a1040,#0f1628);border-bottom:1px solid #1e2d4a;padding:40px 56px}
.logo-tag{font-family:'JetBrains Mono',monospace;font-size:9px;color:#3b82f6;letter-spacing:3px;text-transform:uppercase;margin-bottom:6px;max-width:1400px;margin-left:auto;margin-right:auto}
h1{font-size:26px;font-weight:700;color:#fff;margin-bottom:10px;max-width:1400px;margin-left:auto;margin-right:auto}
.hmeta{display:flex;gap:24px;flex-wrap:wrap;max-width:1400px;margin:0 auto}
.hmeta span{font-family:'JetBrains Mono',monospace;font-size:11px;color:#4a607a}
.hmeta b{color:#cdd9f0}
.main{max-width:1400px;margin:0 auto;padding:36px 56px}
.stat-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:14px;margin-bottom:36px}
.stat{background:#0c1120;border:1px solid #1e2d4a;border-radius:10px;padding:18px;text-align:center}
.stat-n{font-family:'JetBrains Mono',monospace;font-size:28px;font-weight:700;line-height:1;margin-bottom:6px;color:#60a5fa}
.stat-lbl{font-size:10px;color:#4a607a;text-transform:uppercase;letter-spacing:1px}
.sec-hd{font-size:11px;font-weight:600;color:#4a607a;text-transform:uppercase;letter-spacing:2px;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #1e2d4a;margin-top:36px}
.table-wrap{background:#0c1120;border:1px solid #1e2d4a;border-radius:12px;overflow:hidden;margin-bottom:12px;overflow-x:auto}
table{width:100%;border-collapse:collapse}
thead{background:#111827}
th{padding:11px 14px;text-align:left;font-size:10px;font-weight:600;color:#4a607a;text-transform:uppercase;letter-spacing:1px;border-bottom:1px solid #1e2d4a;white-space:nowrap}
td{padding:9px 14px;border-bottom:1px solid #1e2d4a;vertical-align:top;font-size:12px}
tr:last-child td{border-bottom:none}
code{font-family:'JetBrains Mono',monospace;font-size:11px;color:#94a3b8}
.badge{display:inline-block;padding:2px 7px;border-radius:4px;font-size:9px;font-weight:700;letter-spacing:.5px;font-family:monospace;color:#fff;background:#1e3a5f;margin-right:4px}
.badge.alive{background:#166534}
.warn-box{background:#3a1e1e;border:1px solid #7f1d1d;border-radius:8px;padding:14px 18px;margin-bottom:20px;font-size:12px;color:#fca5a5}
.footer{border-top:1px solid #1e2d4a;padding:20px 56px;text-align:center;color:#4a607a;font-family:monospace;font-size:11px;margin-top:20px}
.empty{text-align:center;color:#4a607a;padding:24px}
"""


def _esc(v: Any) -> str:
    return html_module.escape(str(v if v is not None else "-"))


def _recon_html_subdomains(subdomains: List[Dict[str, Any]]) -> str:
    if not subdomains:
        return '<div class="empty">발견된 서브도메인 없음</div>'
    rows = []
    for s in subdomains:
        alive_badge = '<span class="badge alive">LIVE</span>' if s["alive"] else '<span class="badge">unresolved</span>'
        src_badges = "".join(f'<span class="badge">{_esc(src)}</span>' for src in s.get("sources", []))
        rows.append(
            f'<tr><td><code>{_esc(s["host"])}</code></td>'
            f'<td>{alive_badge}</td><td>{src_badges or "-"}</td></tr>'
        )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        '<th>호스트</th><th>생존</th><th>발견 소스</th>'
        '</tr></thead><tbody>' + "".join(rows) + '</tbody></table></div>'
    )


def _recon_html_dns(dns_records: Dict[str, Dict[str, List[str]]]) -> str:
    if not dns_records:
        return '<div class="empty">확인된 DNS 레코드 없음</div>'
    rows = []
    for host in sorted(dns_records.keys()):
        for rtype, values in dns_records[host].items():
            for val in values:
                rows.append(
                    f'<tr><td><code>{_esc(host)}</code></td>'
                    f'<td><span class="badge">{_esc(rtype)}</span></td>'
                    f'<td><code>{_esc(val)}</code></td></tr>'
                )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        '<th>호스트</th><th>타입</th><th>값</th>'
        '</tr></thead><tbody>' + "".join(rows) + '</tbody></table></div>'
    )


def _recon_html_certs(certificates: List[Dict[str, Any]]) -> str:
    if not certificates:
        return '<div class="empty">CT 로그 인증서 없음</div>'
    rows = "".join(
        f'<tr><td><code>{_esc(c.get("common_name"))}</code></td>'
        f'<td>{_esc(c.get("issuer"))}</td>'
        f'<td><code>{_esc(c.get("not_before"))}</code></td>'
        f'<td><code>{_esc(c.get("not_after"))}</code></td></tr>'
        for c in certificates
    )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        '<th>Common Name</th><th>발급자</th><th>발급일</th><th>만료일</th>'
        '</tr></thead><tbody>' + rows + '</tbody></table></div>'
    )


def _recon_html_urls(subdomain_urls: Dict[str, List[Dict[str, Any]]]) -> str:
    if not subdomain_urls:
        return '<div class="empty">수집된 URL 없음</div>'
    rows = []
    for host in sorted(subdomain_urls.keys()):
        for entry in subdomain_urls[host]:
            src_badges = "".join(f'<span class="badge">{_esc(s)}</span>' for s in entry.get("sources", []))
            rows.append(
                f'<tr><td><code>{_esc(host)}</code></td>'
                f'<td><code>{_esc(entry["url"])}</code></td>'
                f'<td>{src_badges or "-"}</td></tr>'
            )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        '<th>서브도메인</th><th>URL</th><th>발견 소스</th>'
        '</tr></thead><tbody>' + "".join(rows) + '</tbody></table></div>'
    )


def _recon_html_ports(ports: Dict[str, Any]) -> str:
    if not ports:
        return '<div class="empty">확인된 열린 포트 없음 (InternetDB 데이터 없음 또는 미실행)</div>'
    rows = []
    for ip in sorted(ports.keys(), key=_ip_sort_key):
        info = ports[ip]
        port_list = ", ".join(str(p) for p in info.get("ports", []))
        hostnames = ", ".join(info.get("hostnames", []))
        cpes = ", ".join(info.get("cpes", []))
        vulns = ", ".join(info.get("vulns", []))
        rows.append(
            f'<tr><td><code>{_esc(ip)}</code></td>'
            f'<td><code>{_esc(port_list)}</code></td>'
            f'<td>{_esc(hostnames)}</td>'
            f'<td><code>{_esc(cpes)}</code></td>'
            f'<td style="color:#fb923c">{_esc(vulns)}</td></tr>'
        )
    return (
        '<div class="table-wrap"><table><thead><tr>'
        '<th>IP</th><th>열린 포트</th><th>호스트명</th><th>CPE</th><th>알려진 취약점(CVE)</th>'
        '</tr></thead><tbody>' + "".join(rows) + '</tbody></table></div>'
    )


def generate_recon_html(result: Dict[str, Any], output_dir: str) -> str:
    """정보수집 결과를 HTML 리포트로 생성하여 파일로 저장하고 절대경로를 반환한다."""
    domain = result["domain"]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname_domain = domain.replace(".", "_")
    fpath = os.path.join(os.path.abspath(output_dir), f"recon_{fname_domain}_{ts}.html")

    meta = result["meta"]
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    warn_block = ""
    if result.get("errors"):
        items = "".join(f'<div>⚠ {_esc(e["source"])}: {_esc(e["message"])}</div>' for e in result["errors"])
        warn_block = f'<div class="warn-box">{items}</div>'

    truncated_note = ""
    if meta.get("subdomain_truncated"):
        truncated_note = (
            f'<span style="color:#fb923c"> (DNS 확인 상한 {meta["max_subdomains"]}개로 제한됨)</span>'
        )

    url_truncated_note = ""
    if meta.get("url_truncated"):
        url_truncated_note = (
            f'<span style="color:#fb923c"> (호스트당 {MAX_URLS_PER_HOST}개 / 전체 {MAX_TOTAL_URLS}개 상한으로 제한됨)</span>'
        )

    html = f"""<!DOCTYPE html>
<html lang="ko"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Space Scan Recon – {_esc(domain)}</title>
<style>{_RECON_CSS}</style></head><body>
<div class="header">
  <div class="logo-tag">SPACE SCAN · RECON (PASSIVE OSINT)</div>
  <h1>패시브 정보수집 리포트</h1>
  <div class="hmeta">
    <span>대상 도메인: <b>{_esc(domain)}</b></span>
    <span>일시: <b>{now}</b></span>
    <span>사용 소스: <b>{_esc(", ".join(meta["sources_used"]))}</b></span>
  </div>
</div>
<div class="main">
{warn_block}
<div class="stat-grid">
  <div class="stat"><div class="stat-n">{meta["subdomain_total"]}</div><div class="stat-lbl">서브도메인</div></div>
  <div class="stat"><div class="stat-n">{len(result["dns_records"])}</div><div class="stat-lbl">DNS 생존 확인</div></div>
  <div class="stat"><div class="stat-n">{len(result["certificates"])}</div><div class="stat-lbl">인증서(CT 로그)</div></div>
  <div class="stat"><div class="stat-n">{meta["url_total"]}</div><div class="stat-lbl">수집 URL</div></div>
  <div class="stat"><div class="stat-n">{len(result["ports"])}</div><div class="stat-lbl">포트 확인 IP</div></div>
</div>

<div class="sec-hd">서브도메인{truncated_note}</div>
{_recon_html_subdomains(result["subdomains"])}

<div class="sec-hd">DNS 레코드</div>
{_recon_html_dns(result["dns_records"])}

<div class="sec-hd">인증서 (CT 로그 — crt.sh)</div>
{_recon_html_certs(result["certificates"])}

<div class="sec-hd">수집 URL — 서브도메인별 (Wayback/Common Crawl/urlscan.io/아카이브 robots·sitemap){url_truncated_note}</div>
{_recon_html_urls(result["subdomain_urls"])}

<div class="sec-hd">열린 포트 (Shodan InternetDB — 무키, 사전 수집 데이터)</div>
{_recon_html_ports(result["ports"])}

</div>
<div class="footer">Generated by Space Scan Recon · {now} · 순수 패시브 — 대상에 직접 요청 없음 · For authorized testing only</div>
</body></html>"""

    os.makedirs(output_dir, exist_ok=True)
    with open(fpath, "w", encoding="utf-8") as fh:
        fh.write(html)
    return fpath


# ── 결과 저장: Excel ─────────────────────────────────────────────────────────

def _safe_cell(v: Any) -> Any:
    """엑셀 수식 인젝션 방어 — 위험 prefix 문자열에 ' 추가."""
    if v is None:
        return ""
    if isinstance(v, str) and v.startswith(_FORMULA_PREFIXES):
        return "'" + v
    return v


def save_recon_to_excel(result: Dict[str, Any], output_dir: str) -> str:
    """정보수집 결과를 xlsx로 저장하고 절대경로를 반환한다.

    시트 구성: INFO / Subdomains / DNS / Certificates / SubdomainURLs / Ports
    """
    from openpyxl import Workbook

    domain = result["domain"]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname_domain = domain.replace(".", "_")
    os.makedirs(output_dir, exist_ok=True)
    fpath = os.path.join(os.path.abspath(output_dir), f"recon_{fname_domain}_{ts}.xlsx")

    meta = result["meta"]
    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "INFO"
    info_rows = [
        ("Domain", domain),
        ("Started At", meta["started_at"]),
        ("Finished At", meta["finished_at"]),
        ("Sources Used", ", ".join(meta["sources_used"])),
        ("Subdomain Total", meta["subdomain_total"]),
        ("Subdomain DNS Resolved", meta["subdomain_resolved"]),
        ("Subdomain Truncated", meta["subdomain_truncated"]),
        ("Max Subdomains", meta["max_subdomains"]),
        ("Resolved IP Count", meta["resolved_ip_count"]),
        ("URL Total", meta["url_total"]),
        ("URL Truncated", meta["url_truncated"]),
        ("Errors", "; ".join(f'{e["source"]}: {e["message"]}' for e in result.get("errors", []))),
    ]
    for k, v in info_rows:
        ws_info.append([_safe_cell(k), _safe_cell(v)])

    ws_sub = wb.create_sheet("Subdomains")
    ws_sub.append(["Host", "Alive", "Sources"])
    for s in result["subdomains"]:
        ws_sub.append([_safe_cell(s["host"]), s["alive"], _safe_cell(", ".join(s.get("sources", [])))])

    ws_dns = wb.create_sheet("DNS")
    ws_dns.append(["Host", "Type", "Value"])
    for host in sorted(result["dns_records"].keys()):
        for rtype, values in result["dns_records"][host].items():
            for val in values:
                ws_dns.append([_safe_cell(host), rtype, _safe_cell(val)])

    ws_cert = wb.create_sheet("Certificates")
    ws_cert.append(["Common Name", "Issuer", "Not Before", "Not After"])
    for c in result["certificates"]:
        ws_cert.append([_safe_cell(c.get("common_name")), _safe_cell(c.get("issuer")),
                        _safe_cell(c.get("not_before")), _safe_cell(c.get("not_after"))])

    ws_urls = wb.create_sheet("SubdomainURLs")
    ws_urls.append(["Subdomain", "URL", "Sources"])
    for host in sorted(result["subdomain_urls"].keys()):
        for entry in result["subdomain_urls"][host]:
            ws_urls.append([
                _safe_cell(host),
                _safe_cell(entry["url"]),
                _safe_cell(", ".join(entry.get("sources", []))),
            ])

    ws_ports = wb.create_sheet("Ports")
    ws_ports.append(["IP", "Ports", "Hostnames", "CPEs", "Tags", "Vulns"])
    for ip in sorted(result["ports"].keys(), key=_ip_sort_key):
        info = result["ports"][ip]
        ws_ports.append([
            _safe_cell(ip),
            _safe_cell(", ".join(str(p) for p in info.get("ports", []))),
            _safe_cell(", ".join(info.get("hostnames", []))),
            _safe_cell(", ".join(info.get("cpes", []))),
            _safe_cell(", ".join(info.get("tags", []))),
            _safe_cell(", ".join(info.get("vulns", []))),
        ])

    wb.save(fpath)
    return fpath
