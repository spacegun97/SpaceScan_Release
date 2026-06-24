#!/usr/bin/env python3
"""
excel_merge.py — 다중 엑셀 파일 스키마-합집합 병합 유틸리티

지원 입력: .xlsx / .xlsm (openpyxl), .xls (xlrd), .csv (stdlib csv)
출처 컬럼: 맨 왼쪽에 고정 '출처파일' 컬럼으로 각 행이 온 파일명 기록

알고리즘 (스키마 합집합 UNION ALL):
  - 각 파일·각 시트의 1행을 헤더로 인식
  - 처음 보는 컬럼명 → master_cols 에 추가 (이전 행들은 해당 칸 빈칸 유지)
  - 동일 컬럼명 → 동일 마스터 위치에 값 채움
  - 컬럼 매칭 기준: 정확히 일치 (대소문자·공백 포함)
"""

import csv
import io
import os
import re
from datetime import datetime
from typing import Any, Dict, Generator, List, Optional, Tuple

# 예약 출처 컬럼명 — 항상 첫 번째 컬럼으로 고정
ORIGIN_COL = "출처파일"

# 지원 확장자 집합
SUPPORTED_EXTS = frozenset({".xlsx", ".xlsm", ".xls", ".csv"})

# 엑셀 수식 인젝션 방어 접두사 목록
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

# 파일명 금지 문자 패턴 (Windows)
_INVALID_FILE_CHARS = re.compile(r'[\\/:*?"<>|]')
_RESERVED_FILENAMES = frozenset({
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
})


# ── 유틸 ──────────────────────────────────────────────────────────────────────

def _safe_cell(v: Any) -> Any:
    """엑셀 수식 인젝션 방어 — 위험 prefix 문자열에 ' 추가. 비문자열은 그대로 반환."""
    if v is None:
        return ""
    if isinstance(v, str) and v.startswith(_FORMULA_PREFIXES):
        return "'" + v
    return v


def _ext(filename: str) -> str:
    """소문자 확장자 반환 (예: '.xlsx')."""
    return os.path.splitext(filename)[1].lower()


def _safe_filename(name: str) -> str:
    """파일명 sanitize — path traversal·예약어·금지문자 차단."""
    name = _INVALID_FILE_CHARS.sub("_", name)
    name = name.replace("..", "_")
    name = name.strip(". ")
    if name.upper() in _RESERVED_FILENAMES:
        name = "_" + name
    return name[:80] or "merge"


# ── 포맷별 시트 리더 ─────────────────────────────────────────────────────────

def _iter_xlsx(data: bytes) -> Generator[Tuple[str, List[List[Any]]], None, None]:
    """openpyxl로 .xlsx/.xlsm 전 시트 순회. (시트명, 행 리스트) yield."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            yield ws.title, rows
    finally:
        wb.close()


def _iter_xls(data: bytes) -> Generator[Tuple[str, List[List[Any]]], None, None]:
    """xlrd로 .xls 전 시트 순회. 날짜 셀은 datetime 객체로 변환."""
    import xlrd
    book = xlrd.open_workbook(file_contents=data)
    for sheet in book.sheets():
        rows: List[List[Any]] = []
        for rx in range(sheet.nrows):
            row: List[Any] = []
            for cx in range(sheet.ncols):
                cell = sheet.cell(rx, cx)
                # xlrd.XL_CELL_DATE == 3
                if cell.ctype == 3:
                    try:
                        row.append(xlrd.xldate_as_datetime(cell.value, book.datemode))
                    except Exception:
                        row.append(cell.value)
                else:
                    row.append(cell.value)
            rows.append(row)
        yield sheet.name, rows


def _iter_csv(
    data: bytes, filename: str
) -> Generator[Tuple[str, List[List[Any]]], None, None]:
    """csv 단일 표 읽기. BOM → utf-8-sig → utf-8 → cp949 인코딩 폴백."""
    text: Optional[str] = None
    for enc in ("utf-8-sig", "utf-8", "cp949"):
        try:
            text = data.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            pass
    if text is None:
        text = data.decode("utf-8", errors="replace")

    reader = csv.reader(io.StringIO(text))
    rows = [list(r) for r in reader]
    # 시트 라벨 = 파일명에서 확장자 제거
    label = os.path.splitext(os.path.basename(filename))[0] or filename
    yield label, rows


def iter_sheets(
    filename: str, data: bytes
) -> Generator[Tuple[str, List[List[Any]]], None, None]:
    """파일명+바이트 → 포맷 판별 → (시트명, 행 리스트) yield.

    지원하지 않는 확장자면 ValueError raise.
    """
    ext = _ext(filename)
    if ext in (".xlsx", ".xlsm"):
        yield from _iter_xlsx(data)
    elif ext == ".xls":
        yield from _iter_xls(data)
    elif ext == ".csv":
        yield from _iter_csv(data, filename)
    else:
        raise ValueError(
            f"지원하지 않는 파일 형식: {os.path.basename(filename)!r} "
            f"(허용: .xlsx / .xlsm / .xls / .csv)"
        )


# ── 스키마 합집합 병합 알고리즘 ──────────────────────────────────────────────

def merge_workbooks(
    sources: List[Tuple[str, bytes]]
) -> Dict[str, Any]:
    """다중 소스를 스키마 합집합으로 병합.

    sources: [(파일명, 바이트스트림), ...]

    반환 dict:
      columns:       최종 컬럼 목록 (0번째 = ORIGIN_COL '출처파일' 고정)
      rows:          행 리스트 — 각 행은 {master_col_index: value} dict
      per_file:      파일별 통계 list
      total_rows:    누적 총 행 수
      skipped_files: 처리 실패/미지원 파일명 목록
    """
    # 출처 컬럼을 0번째 고정 예약
    master_cols: List[str] = [ORIGIN_COL]
    col_index: Dict[str, int] = {ORIGIN_COL: 0}
    rows_out: List[Dict[int, Any]] = []
    per_file: List[Dict[str, Any]] = []
    skipped_files: List[str] = []
    empty_col_counter = [0]  # 빈 헤더 셀 전역 고유명 카운터

    for filename, data in sources:
        file_stat: Dict[str, Any] = {
            "name": filename,
            "sheets_read": 0,
            "rows_added": 0,
            "new_columns": [],
            "skipped": False,
            "error": None,
        }

        try:
            for _sheet_label, raw_rows in iter_sheets(filename, data):
                if not raw_rows:
                    continue  # 빈 시트 스킵

                # 1행 = 헤더
                header_raw = raw_rows[0]
                if not header_raw or all(
                    v is None or str(v).strip() == "" for v in header_raw
                ):
                    continue  # 완전 빈 헤더 시트 스킵

                # 헤더 정규화 + 빈 헤더 고유명 부여 + 시트 내 중복 접미사 분리
                seen_in_sheet: Dict[str, int] = {}
                local_col_names: List[str] = []

                for raw_val in header_raw:
                    name = "" if raw_val is None else str(raw_val)

                    if name == "":
                        # 빈 헤더 → 전역 고유 임시명 부여
                        while True:
                            candidate = f"(빈컬럼_{empty_col_counter[0]})"
                            empty_col_counter[0] += 1
                            if candidate not in seen_in_sheet:
                                break
                        name = candidate

                    # 이 시트 내 중복 컬럼 처리 — 두 번째부터 .1 .2 ... 접미사
                    if name in seen_in_sheet:
                        seen_in_sheet[name] += 1
                        name = f"{name}.{seen_in_sheet[name]}"
                    else:
                        seen_in_sheet[name] = 0

                    local_col_names.append(name)

                # 예약 출처 컬럼명 '출처파일' 충돌 방어
                # 데이터에 '출처파일' 컬럼이 있으면 '출처파일.1'로 변환
                safe_local: List[str] = [
                    f"{ORIGIN_COL}.1" if c == ORIGIN_COL else c
                    for c in local_col_names
                ]

                # master_cols 에 신규 컬럼 등록 + 로컬→마스터 인덱스 매핑 테이블 구성
                local_map: List[int] = []
                for cname in safe_local:
                    if cname not in col_index:
                        col_index[cname] = len(master_cols)
                        master_cols.append(cname)
                        file_stat["new_columns"].append(cname)
                    local_map.append(col_index[cname])

                # 데이터 행 처리 (헤더 이후)
                for raw_row in raw_rows[1:]:
                    if not raw_row or all(
                        v is None or str(v).strip() == "" for v in raw_row
                    ):
                        continue  # 완전 빈 행 스킵

                    # 0번째 = 출처파일, 나머지는 로컬 매핑으로 올바른 마스터 위치에 배치
                    rec: Dict[int, Any] = {0: filename}
                    for ci, val in enumerate(raw_row):
                        if ci < len(local_map):
                            rec[local_map[ci]] = val
                    rows_out.append(rec)
                    file_stat["rows_added"] += 1

                file_stat["sheets_read"] += 1

        except ValueError as e:
            # 미지원 확장자
            file_stat["skipped"] = True
            file_stat["error"] = str(e)
            skipped_files.append(filename)
        except Exception as e:
            # 파일 손상 등 기타 오류
            file_stat["skipped"] = True
            file_stat["error"] = f"파일 읽기 오류: {e}"
            skipped_files.append(filename)

        per_file.append(file_stat)

    return {
        "columns": master_cols,
        "rows": rows_out,
        "per_file": per_file,
        "total_rows": len(rows_out),
        "skipped_files": skipped_files,
    }


# ── 결과 저장 ────────────────────────────────────────────────────────────────

def save_merged(
    result: Dict[str, Any],
    output_dir: str,
    out_name: str,
) -> str:
    """병합 결과를 단일 'Merged' 시트 xlsx로 저장하고 절대경로 반환.

    출력 파일명: merge_<out_name>_<YYYYMMDD_HHMMSS>.xlsx
    빠진 컬럼은 빈칸으로 채워 모든 행이 동일한 너비를 유지한다.
    """
    from openpyxl import Workbook

    os.makedirs(output_dir, exist_ok=True)
    safe_name = _safe_filename(out_name)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_filename = f"merge_{safe_name}_{ts}.xlsx"
    filepath = os.path.join(output_dir, out_filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Merged"

    columns: List[str] = result["columns"]
    rows: List[Dict[int, Any]] = result["rows"]
    col_count = len(columns)

    # 헤더 행 기록
    ws.append([_safe_cell(c) for c in columns])

    # 데이터 행 기록 — rec에 없는 컬럼 인덱스는 빈칸("")으로 채움
    for rec in rows:
        ws.append([_safe_cell(rec.get(i)) for i in range(col_count)])

    wb.save(filepath)
    return os.path.abspath(filepath)
